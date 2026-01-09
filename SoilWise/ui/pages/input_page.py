"""
SoilWise/ui/pages/input_page.py
Enhanced Soil data input page with complete evaluation integration
Based on Escomen et al. 2024 methodology and Square Root Method
"""

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QScrollArea,
                                QFrame, QGridLayout, QLineEdit, QComboBox, QGroupBox,
                                QDoubleSpinBox, QMessageBox, QFileDialog, QPushButton, QApplication)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont, QColor, QPalette
from PySide6.QtWidgets import QGraphicsDropShadowEffect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from database.db_manager import get_database


# Import evaluation engine
try:
    from knowledge_base.evaluation import SuitabilityEvaluator
    EVALUATOR_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Warning: Could not import SuitabilityEvaluator: {e}")
    EVALUATOR_AVAILABLE = False


class EnhancedButton(QPushButton):
    """Enhanced button with modern styling"""
    def __init__(self, text, icon="", primary=False, parent=None):
        super().__init__(parent)
        self.setText(f"{icon} {text}" if icon else text)
        self.primary = primary
        self.setCursor(Qt.PointingHandCursor)
        self.setMinimumHeight(48)
        self.apply_style()

    def apply_style(self):
        """Apply modern agricultural styling"""
        if self.primary:
            self.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #7d9d7f, stop:1 #6b8a6d);
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-size: 15px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #8ab08c, stop:1 #7d9d7f);
                }
                QPushButton:pressed {
                    background: #6b8a6d;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background: white;
                    color: #5a7a5c;
                    border: 2px solid #d4e4d4;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-size: 14px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background: #f5f9f5;
                    border-color: #7d9d7f;
                }
                QPushButton:pressed {
                    background: #eaf3ea;
                }
            """)


class InputPage(QWidget):
    """Enhanced Soil data input page with complete evaluation integration"""
    
    data_saved = Signal(int)
    evaluation_complete = Signal(dict)  # Emits evaluation results for navigation

    def __init__(self, parent=None):
        super().__init__(parent)

        self.current_soil_crop = None
        self.soil_inputs = {}
        self.climate_inputs = {}
        self.crop_input = None
        self.season_input = None
        self.season_label = None
        self.seasonal_info = None

        
        # Define seasonal crops based on Escomen et al. Table 7
        self.seasonal_crops = {
            "Cabbage", "Carrots", "Maize", 
            "Sorghum", "Sweet Potato", "Tomato"
        }
        

        # ✅ NEW: Cache for dropdown options to prevent loss on clear
        self._cached_site_names = [
            "Select barangay...", "Aposong", "Bagoaingud", "Bangco (Pob.)",
            "Bansayan", "Basak", "Bobo", "Bualan", "Bubong Ilian",
            "Bubong Tawa-an", "Bubonga Mamaanun", "Gacap", "Ilian",
            "Ilian Poblacion", "Kalanganan", "Katumbacan", "Lininding",
            "Lumbaca Mamaan", "Mamaanun", "Mentring", "Olango", "Palacat",
            "Palao", "Paling", "Pantaon", "Pantar", "Paridi", "Pindolonan",
            "Radapan", "Radapan Poblacion", "Rantian", "Sapingit", "Talao",
            "Tambo", "Tapocan", "Taporug", "Tawaan", "Udalo"
        ]

        self._cached_flooding_options = [
            "Select flooding class...",
            "Fo - No flooding",
            "F1 - Occasional flooding",
            "F2 - Frequent flooding",
            "F3 - Very frequent flooding",
            "F1+ - Severe flooding"
        ]

        self._cached_default_drainage_options = [
            "Select drainage class...",
            "good - Well drained",
            "good_gw_over_150 - Good drainage, groundwater >150 cm",
            "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
            "moderate - Moderately drained",
            "imperfect - Imperfectly drained",
            "poor_drainable - Poorly drained but drainable",
            "poor_not_drainable - Poorly drained, not drainable"
        ]

        self._cached_default_texture_options = [
            "Select texture...",
            "C - Clay",
            "CL - Clay Loam",
            "L - Loam",
            "S - Sand"
        ]

        self._current_drainage_options = self._cached_default_drainage_options.copy()
        self._current_texture_options = self._cached_default_texture_options.copy()

        # Initialize evaluation engine
        self.evaluator = None
        if EVALUATOR_AVAILABLE:
            try:
                self.evaluator = SuitabilityEvaluator()
                print("✅ Evaluation engine initialized successfully")
            except Exception as e:
                print(f"⚠️ Warning: Could not initialize evaluator: {e}")
        else:
            print("⚠️ Warning: Evaluation engine not available")
        
        self.init_ui()

        # Initialize database connection
        try:
            self.db = get_database()
            print("Database connected in Input Page")
        except Exception as e:
            print(f"Database connection failed: {e}")
            self.db = None


    def init_ui(self):
        """Initialize enhanced user interface"""
        self.setStyleSheet("background-color: #fafcfa;")
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: #fafcfa; }")
        
        container = QWidget()
        container.setStyleSheet("background-color: #fafcfa;")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(24)
        
        # Page header
        header = self.create_header()
        layout.addWidget(header)
        
        # Import/Export card
        layout.addWidget(self.create_import_export_card())
        
        # Crop selection with seasonal support
        layout.addWidget(self.create_crop_selection_group())
        
        # Location group
        layout.addWidget(self.create_location_group())
        
        # Soil properties with subcategories
        layout.addWidget(self.create_soil_properties_group())
        
        # Climate group
        layout.addWidget(self.create_climate_group())
        
        # Action buttons
        layout.addLayout(self.create_action_buttons())
        
        # Analysis section
        layout.addWidget(self.create_analysis_card())
        
        layout.addStretch()
        
        scroll.setWidget(container)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def create_header(self):
        """Create page header"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 16)
        layout.setSpacing(8)
        
        title = QLabel("Soil Data Input")
        title.setFont(QFont("Georgia", 28, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")
        
        desc = QLabel("Enter detailed soil and landscape characteristics for crop suitability evaluation")
        desc.setFont(QFont("Segoe UI", 14))
        desc.setStyleSheet("color: #6a8a6c; line-height: 1.5;")
        desc.setWordWrap(True)
        
        layout.addWidget(title)
        layout.addWidget(desc)
        
        return widget

    def create_import_export_card(self):
        """Create enhanced import/export card"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)
        
        layout = QHBoxLayout(card)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)
        
        label = QLabel("Data Import/Export")
        label.setFont(QFont("Georgia", 16, QFont.Bold))
        label.setStyleSheet("color: #3d5a3f;")
        
        template_btn = EnhancedButton("Download Template", "")
        template_btn.clicked.connect(self.download_template)
        
        import_btn = EnhancedButton("Import Excel", "↓")
        import_btn.clicked.connect(self.import_excel)
        
        export_btn = EnhancedButton("Export Excel", "↑")
        export_btn.clicked.connect(self.export_excel)
        
        layout.addWidget(label)
        layout.addStretch()
        layout.addWidget(template_btn)
        layout.addWidget(import_btn)
        layout.addWidget(export_btn)
        
        return card

    def create_crop_selection_group(self):
        """Create crop selection group with seasonal support"""
        group = QGroupBox()
        group.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        group.setGraphicsEffect(shadow)
        
        title_label = QLabel("⚘ Crop Selection")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        
        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)
        layout.addWidget(title_label)
        
        grid = QGridLayout()
        grid.setSpacing(20)
        grid.setColumnStretch(1, 1)
        
        # Crop selection dropdown
        crop_label = QLabel("Select Crop:")
        crop_label.setFont(QFont("Segoe UI", 13, QFont.DemiBold))
        crop_label.setStyleSheet("color: #4a6a4c;")
        grid.addWidget(crop_label, 0, 0)
        
        self.crop_input = QComboBox()
        self.crop_input.addItems([
            "Select a crop...",
            "Arabica Coffee", "Banana", "Cabbage", "Carrots", "Cocoa",
            "Maize", "Oil Palm", "Pineapple", "Robusta Coffee",
            "Sorghum", "Sugarcane", "Sweet Potato", "Tomato"
        ])
        self.crop_input.setMinimumHeight(44)
        self.crop_input.setStyleSheet("""
            QComboBox {
                background: #f9fbf9;
                border: 2px solid #e0ede0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
            }
            QComboBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 5px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
                margin-right: 8px;
            }
        """)
        self.crop_input.currentTextChanged.connect(self.on_crop_changed)
        grid.addWidget(self.crop_input, 0, 1)
        
        # Season selection
        self.season_label = QLabel("Growing Season:")
        self.season_label.setFont(QFont("Segoe UI", 13, QFont.DemiBold))
        self.season_label.setStyleSheet("color: #4a6a4c;")
        self.season_label.setVisible(False)
        grid.addWidget(self.season_label, 1, 0)
        
        self.season_input = QComboBox()
        self.season_input.addItems([
            "Select season...",
            "January - April (Dry Season)",
            "May - August (Wet Season)",
            "September - December (Cool Season)"
        ])
        self.season_input.setMinimumHeight(44)
        self.season_input.setStyleSheet("""
            QComboBox {
                background: #fff8dc;
                border: 2px solid #e0ede0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
            }
            QComboBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 5px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
                margin-right: 8px;
            }
        """)
        self.season_input.setVisible(False)
        grid.addWidget(self.season_input, 1, 1)
        
        self.seasonal_info = QLabel("⚠️ This crop has different suitability across growing seasons")
        self.seasonal_info.setFont(QFont("Segoe UI", 11, QFont.Normal))
        self.seasonal_info.setStyleSheet("color: #c87b00; font-style: italic;")
        self.seasonal_info.setWordWrap(True)
        self.seasonal_info.setVisible(False)
        grid.addWidget(self.seasonal_info, 2, 0, 1, 2)
        
        layout.addLayout(grid)
        group.setLayout(layout)
        
        return group

    def on_crop_changed(self, crop_name):
        """Show/hide season selector, update drainage and texture options, and guard reusing soil across crops."""
        
        is_seasonal = crop_name in self.seasonal_crops
        
        if self.season_label:
            self.season_label.setVisible(is_seasonal)
        if self.season_input:
            self.season_input.setVisible(is_seasonal)
        if self.seasonal_info:
            self.seasonal_info.setVisible(is_seasonal)
        
        if not is_seasonal and self.season_input:
            self.season_input.setCurrentIndex(0)
        
        # === UPDATE BOTH DRAINAGE AND TEXTURE OPTIONS ===
        if crop_name != "Select a crop...":
            self.update_drainage_options(crop_name)
            self.update_texture_options(crop_name)  # ← ADD THIS LINE
        # ================================================
        
        if (self.current_soil_crop and crop_name and 
            crop_name != "Select a crop..." and crop_name != self.current_soil_crop):
            
            reply = QMessageBox.question(
                self, 
                "Use existing soil data?",
                f"The current soil inputs entered were for {self.current_soil_crop}. "
                f"It is recommended to use another set of soil data. "
                f"Start a new one instead?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.No:
                self.clear_form()
            else:
                self.current_soil_crop = crop_name
        elif crop_name != "Select a crop...":
            self.current_soil_crop = crop_name


    def get_selected_season_code(self):
        """Convert UI season text to API season code"""
        if not self.season_input:
            return None
            
        season_text = self.season_input.currentText()
        season_mapping = {
            "January - April (Dry Season)": "january_april",
            "May - August (Wet Season)": "may_august",
            "September - December (Cool Season)": "september_december"
        }
        return season_mapping.get(season_text)

    def create_location_group(self):
        """Create location information group"""
        group = QGroupBox()
        group.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        group.setGraphicsEffect(shadow)
        
        title_label = QLabel("◉ Location Information")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        
        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)
        layout.addWidget(title_label)
        
        grid = QGridLayout()
        grid.setSpacing(20)
        grid.setColumnStretch(1, 1)
        
        site_label = QLabel("Site Name (Barangay):")
        site_label.setFont(QFont("Segoe UI", 13, QFont.DemiBold))
        site_label.setStyleSheet("color: #4a6a4c;")
        grid.addWidget(site_label, 0, 0)
        
        self.site_input = QComboBox()
        self.site_input.addItems([
            "Select barangay...",
            "Aposong", "Bagoaingud", "Bangco (Pob.)", "Bansayan",
            "Basak", "Bobo", "Bualan", "Bubong Ilian",
            "Bubong Tawa-an", "Bubonga Mamaanun", "Gacap", "Ilian",
            "Ilian Poblacion", "Kalanganan", "Katumbacan", "Lininding",
            "Lumbaca Mamaan", "Mamaanun", "Mentring", "Olango",
            "Palacat", "Palao", "Paling", "Pantaon", "Pantar",
            "Paridi", "Pindolonan", "Radapan", "Radapan Poblacion",
            "Rantian", "Sapingit", "Talao", "Tambo", "Tapocan",
            "Taporug", "Tawaan", "Udalo"
        ])
        self.site_input.setMinimumHeight(44)
        self.site_input.setStyleSheet("""
            QComboBox {
                background: #f9fbf9;
                border: 2px solid #e0ede0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
            }
            QComboBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 5px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
                margin-right: 8px;
            }
        """)
        grid.addWidget(self.site_input, 0, 1)
        
        layout.addLayout(grid)
        group.setLayout(layout)
        
        return group

    def create_soil_properties_group(self):
        """Create comprehensive soil properties group with subcategories"""
        group = QGroupBox()
        group.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        group.setGraphicsEffect(shadow)
        
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(28, 32, 28, 28)
        main_layout.setSpacing(32)
        
        title_label = QLabel("◉ Soil Properties")
        title_label.setFont(QFont("Georgia", 18, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        main_layout.addWidget(title_label)
        
        main_layout.addWidget(self.create_subsection("Topography", [
            ("Slope (%):", "slope", 0, 100, 0, 0.1),
        ]))
        
        main_layout.addWidget(self.create_wetness_subsection())
        
        main_layout.addWidget(self.create_subsection("Physical Soil Characteristics", [
            ("Coarse Fragments (vol %):", "coarse_fragments", 0, 100, 0, 1.0),
            ("Soil Depth (cm):", "soil_depth", 0, 300, 0, 1.0),
            ("CaCO₃ (%):", "caco3", 0, 100, 0, 0.1),
            ("Gypsum (%):", "gypsum", 0, 100, 0, 0.1),
        ], include_texture=True))
        
        # ✅ MODIFIED: Added Sum of Basic Cations field
        main_layout.addWidget(self.create_subsection("Soil Fertility Characteristics", [
            ("Apparent CEC (cmol/kg clay):", "cec", 0, 200, 0, 0.1),
            ("Sum of Basic Cations (cmol/kg):", "sum_basic_cations", 0, 100, 0, 0.1),  # NEW FIELD
            ("Base Saturation (%):", "base_saturation", 0, 100, 0, 0.1),
            ("pH (H₂O):", "ph", 0, 14, 0, 0.1),
            ("Organic Carbon (%):", "organic_carbon", 0, 10, 0, 0.1),
        ]))
        
        main_layout.addWidget(self.create_subsection("Salinity and Alkalinity", [
            ("ECe (dS/m):", "ece", 0, 20, 0, 0.01),
            ("ESP (%):", "esp", 0, 100, 0, 0.01),
        ]))
        
        group.setLayout(main_layout)
        return group

    def create_subsection(self, title, fields, include_texture=False):
        """Create a subsection with fields"""
        container = QFrame()
        container.setStyleSheet("""
            QFrame {
                background: #f9fbf9;
                border-radius: 8px;
                border: 1px solid #e8f1e8;
                padding: 16px;
            }
        """)
        
        layout = QVBoxLayout(container)
        layout.setSpacing(16)
        
        subtitle = QLabel(title)
        subtitle.setFont(QFont("Segoe UI", 14, QFont.DemiBold))
        subtitle.setStyleSheet("color: #5a7a5c; background: transparent; border: none; padding: 0;")
        layout.addWidget(subtitle)
        
        grid = QGridLayout()
        grid.setSpacing(16)
        grid.setColumnStretch(1, 1)
        
        spinbox_style = """
            QDoubleSpinBox {
                background: white;
                border: 2px solid #e0ede0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
                min-height: 28px;
            }
            QDoubleSpinBox:focus {
                border-color: #7d9d7f;
            }
            QDoubleSpinBox::up-button {
                background: #e8f3e8;
                border: none;
                border-top-right-radius: 4px;
                width: 20px;
            }
            QDoubleSpinBox::up-button:hover {
                background: #d4e4d4;
            }
            QDoubleSpinBox::down-button {
                background: #e8f3e8;
                border: none;
                border-bottom-right-radius: 4px;
                width: 20px;
            }
            QDoubleSpinBox::down-button:hover {
                background: #d4e4d4;
            }
            QDoubleSpinBox::up-arrow {
                image: none;
                border: 4px solid transparent;
                border-bottom: 6px solid #3d5a3f;
                width: 0;
                height: 0;
            }
            QDoubleSpinBox::down-arrow {
                image: none;
                border: 4px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
            }
        """
        
        for i, field_data in enumerate(fields):
            label_text, key, min_val, max_val, default, step = field_data
            
            label = QLabel(label_text)
            label.setFont(QFont("Segoe UI", 12, QFont.Medium))
            label.setStyleSheet("color: #4a6a4c; background: transparent; border: none;")
            grid.addWidget(label, i, 0)
            
            spinbox = QDoubleSpinBox()
            spinbox.setRange(min_val, max_val)
            spinbox.setValue(default)
            spinbox.setSingleStep(step)
            spinbox.setMinimumHeight(40)
            spinbox.setStyleSheet(spinbox_style)
            
            self.soil_inputs[key] = spinbox
            grid.addWidget(spinbox, i, 1)
        
        if include_texture:
            texture_label = QLabel("Soil Texture")
            texture_label.setFont(QFont("Segoe UI", 12, QFont.Medium))
            texture_label.setStyleSheet("color: #4a6a4c; background: transparent; border: none;")
            grid.addWidget(texture_label, len(fields), 0)
            
            self.texture_input = QComboBox()
            # Start with default options - will be updated by crop selection
            self.texture_input.addItems([
                "Select texture...",
                "C - Clay",
                "CL - Clay Loam",
                "L - Loam",
                "S - Sand"
            ])
            self.texture_input.setMinimumHeight(40)
            self.texture_input.setStyleSheet("""
                QComboBox {
                    background: white;
                    border: 2px solid #e0ede0;
                    border-radius: 6px;
                    padding: 8px 12px;
                    font-size: 14px;
                    color: #3d5a3f;
                }
                QComboBox:focus {
                    border-color: #7d9d7f;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 30px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border: 5px solid transparent;
                    border-top: 6px solid #3d5a3f;
                    width: 0;
                    height: 0;
                    margin-right: 8px;
                }
            """)
            grid.addWidget(self.texture_input, len(fields), 1)
        
        layout.addLayout(grid)
        return container


    def create_wetness_subsection(self):
        """
        Create wetness subsection with flooding and drainage.
        Drainage options dynamically update based on selected crop.
        """
        container = QFrame()
        container.setStyleSheet("""
            QFrame {
                background: #f9fbf9;
                border-radius: 8px;
                border: 1px solid #e8f1e8;
                padding: 16px;
            }
        """)
        
        layout = QVBoxLayout(container)
        layout.setSpacing(16)
        
        subtitle = QLabel("Wetness")
        subtitle.setFont(QFont("Segoe UI", 14, QFont.DemiBold))
        subtitle.setStyleSheet("color: #5a7a5c; background: transparent; border: none; padding: 0;")
        layout.addWidget(subtitle)
        
        grid = QGridLayout()
        grid.setSpacing(16)
        grid.setColumnStretch(1, 1)
        
        combo_style = """
            QComboBox {
                background: white;
                border: 2px solid #e0ede0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 13px;
                color: #3d5a3f;
            }
            QComboBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 5px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
                margin-right: 8px;
            }
        """
        
        # FLOODING
        flooding_label = QLabel("Flooding")
        flooding_label.setFont(QFont("Segoe UI", 12, QFont.Medium))
        flooding_label.setStyleSheet("color: #4a6a4c; background: transparent; border: none;")
        grid.addWidget(flooding_label, 0, 0)
        
        self.flooding_input = QComboBox()
        self.flooding_input.addItems([
            "Select flooding class...",
            "Fo - No flooding",
            "F1 - Occasional flooding",
            "F2 - Frequent flooding", 
            "F3 - Very frequent flooding",
            "F3+ - Severe flooding"
        ])
        self.flooding_input.setMinimumHeight(40)
        self.flooding_input.setStyleSheet(combo_style)
        grid.addWidget(self.flooding_input, 0, 1)
        
        # DRAINAGE
        drainage_label = QLabel("Drainage")
        drainage_label.setFont(QFont("Segoe UI", 12, QFont.Medium))
        drainage_label.setStyleSheet("color: #4a6a4c; background: transparent; border: none;")
        grid.addWidget(drainage_label, 1, 0)
        
        self.drainage_input = QComboBox()
        self.drainage_input.addItems([
            "Select drainage class...",
            "good - Well drained",
            "good_gw_over_150 - Good drainage, groundwater >150 cm",
            "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
            "moderate - Moderately drained",
            "moder. - Moderately drained (alternate)",
            "imperfect - Imperfectly drained",
            "poor_aeric - Poorly drained, aeric conditions",
            "poor_drainable - Poorly drained but drainable",
            "poor_not_drainable - Poorly drained, not drainable",
            "poor - Poorly drained",
            "very_poor - Very poorly drained"
        ])
        self.drainage_input.setMinimumHeight(40)
        self.drainage_input.setStyleSheet(combo_style)
        grid.addWidget(self.drainage_input, 1, 1)
        
        layout.addLayout(grid)
        return container


    def create_climate_group(self):
        """Create climate characteristics group"""
        group = QGroupBox()
        group.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        group.setGraphicsEffect(shadow)
        
        title_label = QLabel("◎ Climate Characteristics")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        
        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)
        layout.addWidget(title_label)
        
        grid = QGridLayout()
        grid.setSpacing(20)
        grid.setColumnStretch(1, 1)
        
        properties = [
            ("Average Temperature (°C):", "temperature", 0, 50, 0, 0.1),
            ("Annual Rainfall (mm):", "rainfall", 0, 5000, 0, 10),
            ("Humidity (%):", "humidity", 0, 100, 0, 1),
        ]
        
        spinbox_style = """
            QDoubleSpinBox {
                background: #f9fbf9;
                border: 2px solid #e0ede0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
                min-height: 28px;
            }
            QDoubleSpinBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QDoubleSpinBox::up-button {
                background: #e8f3e8;
                border: none;
                border-top-right-radius: 6px;
                width: 20px;
            }
            QDoubleSpinBox::up-button:hover {
                background: #d4e4d4;
            }
            QDoubleSpinBox::down-button {
                background: #e8f3e8;
                border: none;
                border-bottom-right-radius: 6px;
                width: 20px;
            }
            QDoubleSpinBox::down-button:hover {
                background: #d4e4d4;
            }
            QDoubleSpinBox::up-arrow {
                image: none;
                border: 4px solid transparent;
                border-bottom: 6px solid #3d5a3f;
                width: 0;
                height: 0;
            }
            QDoubleSpinBox::down-arrow {
                image: none;
                border: 4px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
            }
        """
        
        for i, (label_text, key, min_val, max_val, default, step) in enumerate(properties):
            label = QLabel(label_text)
            label.setFont(QFont("Segoe UI", 13, QFont.DemiBold))
            label.setStyleSheet("color: #4a6a4c;")
            grid.addWidget(label, i, 0)
            
            spinbox = QDoubleSpinBox()
            spinbox.setRange(min_val, max_val)
            spinbox.setValue(default)
            spinbox.setSingleStep(step)
            spinbox.setMinimumHeight(44)
            spinbox.setStyleSheet(spinbox_style)
            
            self.climate_inputs[key] = spinbox
            grid.addWidget(spinbox, i, 1)
        
        layout.addLayout(grid)
        group.setLayout(layout)
        
        return group

    def create_action_buttons(self):
        """Create action buttons"""
        layout = QHBoxLayout()
        layout.setSpacing(16)
        layout.addStretch()

        btn_clear = EnhancedButton("Clear Form", "↻")
        btn_clear.setMinimumWidth(160)
        btn_clear.clicked.connect(self.clear_form)

        btn_save = EnhancedButton("Save Data", "", primary=True)
        btn_save.setMinimumWidth(160)
        btn_save.setMinimumHeight(52)
        btn_save.clicked.connect(self.save_data)

        layout.addWidget(btn_clear)
        layout.addWidget(btn_save)
        return layout
    
    def clear_form(self):
        """
        Clear all form inputs while preserving dropdown options.
        ✅ FIXED: Now uses cached lists to restore dropdowns after clearing.
        """
        # Clear climate numeric inputs
        for widget in self.climate_inputs.values():
            if isinstance(widget, QDoubleSpinBox):
                widget.setValue(0.0)

        # Clear soil numeric inputs
        for widget in self.soil_inputs.values():
            if isinstance(widget, QDoubleSpinBox):
                widget.setValue(0.0)

        # ✅ FIXED: Reset dropdowns to first item (index 0) instead of clearing items
        if hasattr(self, 'texture_input') and self.texture_input:
            # Restore cached texture options if needed
            if self.texture_input.count() == 0:
                self.texture_input.addItems(self._current_texture_options)
            self.texture_input.setCurrentIndex(0)

        if hasattr(self, 'drainage_input') and self.drainage_input:
            # Restore cached drainage options if needed
            if self.drainage_input.count() == 0:
                self.drainage_input.addItems(self._current_drainage_options)
            self.drainage_input.setCurrentIndex(0)

        if hasattr(self, 'flooding_input') and self.flooding_input:
            # Restore cached flooding options if needed
            if self.flooding_input.count() == 0:
                self.flooding_input.addItems(self._cached_flooding_options)
            self.flooding_input.setCurrentIndex(0)

        # ✅ FIXED: Reset site selection while preserving the list
        if hasattr(self, 'site_input') and self.site_input:
            # Restore cached site names if needed
            if self.site_input.count() == 0:
                self.site_input.addItems(self._cached_site_names)
            self.site_input.setCurrentIndex(0)

        # Reset crop and season selections
        if self.cropinput:
            self.crop_input.setCurrentIndex(0)

        if self.season_input:
            self.season_input.setCurrentIndex(0)

        # Reset soil-crop association
        self.current_soil_crop = None

        print("✅ Form cleared successfully - all dropdowns preserved")


    def create_analysis_card(self):
        """Create enhanced analysis section"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #e8f3e8, stop:1 #f0f7f0);
                border-radius: 12px;
                border: none;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(12)
        
        title = QLabel("Ready to Analyze?")
        title.setFont(QFont("Georgia", 20, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")
        
        desc = QLabel(
            "Once you've entered soil and climate data, "
            "click the button below to run the crop suitability analysis."
        )
        desc.setFont(QFont("Segoe UI", 14))
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #5a7a5c; margin: 8px 0 20px 0;")
        
        btn_analyze = EnhancedButton("Run Analysis", "", primary=True)
        btn_analyze.setMinimumHeight(64)
        btn_analyze.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #7d9d7f, stop:1 #6b8a6d);
                color: white;
                border: none;
                border-radius: 10px;
                padding: 16px 32px;
                font-size: 16px;
                font-weight: 700;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #8ab08c, stop:1 #7d9d7f);
            }
            QPushButton:pressed {
                background: #6b8a6d;
            }
        """)
        btn_analyze.clicked.connect(self.run_analysis)
        
        layout.addWidget(title)
        layout.addWidget(desc)
        layout.addWidget(btn_analyze)
        
        return card

    def get_texture_code(self):
        """Extract USDA texture code from dropdown"""
        if not hasattr(self, 'texture_input'):
            return ""
        
        texture_text = self.texture_input.currentText()
        if texture_text == "Select texture..." or not texture_text:
            return ""
        
        code = texture_text.split(" - ")[0].strip()
        return code

    def get_flooding_code(self):
        """Extract flooding code from dropdown"""
        if not hasattr(self, 'flooding_input'):
            return ""
        
        flooding_text = self.flooding_input.currentText()
        if flooding_text == "Select flooding class..." or not flooding_text:
            return ""
        
        code = flooding_text.split(" - ")[0].strip()
        return code

    def get_drainage_code(self):
        """Extract drainage code from dropdown"""
        if not hasattr(self, 'drainage_input'):
            return ""
        
        drainage_text = self.drainage_input.currentText()
        if drainage_text == "Select drainage class..." or not drainage_text:
            return ""
        
        code = drainage_text.split(" - ")[0].strip()
        return code
    
    def update_drainage_options(self, crop_name):
        """Update drainage dropdown based on selected crop from Land Evaluation Book Part 3."""
        drainage_options = {
            'Arabica Coffee': [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater >150 cm",
                "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            'Banana': [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater >150 cm",
                "moderate - Moderately drained",
                "moder. - Moderately drained (alternate)",
                "imperfect - Imperfectly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            'Cocoa': [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater >150 cm",
                "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor - Poorly drained",
                "very_poor - Very poorly drained"
            ],
            'Cabbage': [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater >150 cm",
                "moderate - Moderately drained",
                "moder. - Moderately drained (alternate)",
                "imperfect - Imperfectly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            'Carrots': [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater >150 cm",
                "moderate - Moderately drained",
                "moder. - Moderately drained (alternate)",
                "imperfect - Imperfectly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            "Robusta Coffee": [
                "Select drainage class...",
                "good - Well drained",
                "good_gw_over_150 - Good drainage, groundwater > 150 cm",
                "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor - Poorly drained",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            "Maize": [
                "Select drainage class...",
                "good - Well drained",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor - Poorly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            "Oil Palm": [
                "Select drainage class...",
                "good - Well drained",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor - Poorly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            "Pineapple": [
                "Select drainage class...",
                "good - Well drained",
                "goodgwover150 - Well drained (GW >150cm)",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor - Poorly drained",
                "pooraeric - Poorly drained, aeric conditions",
                "poordrainable - Poorly drained but drainable",
                "poornotdrainable - Poorly drained, not drainable"
            ],
            "Sorghum": [
                "Select drainage class...",
                "good - Well drained",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
                ],
            "Sugarcane": [
                "Select drainage class...",
                "good - Well drained",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ],
            "Sweet Potato": [
                "Select drainage class...",
                "good - Well drained",
                "moderate - Moderately drained",
                "imperfect - Imperfectly drained",
                "poor_and_aeric - Poorly drained, aeric conditions",
                "poor_drainable - Poorly drained but drainable",
                "poor_not_drainable - Poorly drained, not drainable"
            ]

        }
        
        default_options = [
            "Select drainage class...",
            "good - Well drained",
            "good_gw_over_150 - Good drainage, groundwater >150 cm",
            "good_gw_100_150 - Good drainage, groundwater 100-150 cm",
            "moderate - Moderately drained",
            "imperfect - Imperfectly drained",
            "poor_drainable - Poorly drained but drainable",
            "poor_not_drainable - Poorly drained, not drainable"
        ]
        
        if not hasattr(self, 'drainage_input') or not self.drainage_input:
            return
        
        current_code = self.get_drainage_code()
        options = drainage_options.get(crop_name, default_options)

        # ✅ NEW: Cache the current drainage options for this crop
        self._current_drainage_options = options.copy()
        
        self.drainage_input.clear()
        self.drainage_input.addItems(options)
        
        if current_code:
            for i in range(self.drainage_input.count()):
                item_text = self.drainage_input.itemText(i)
                if item_text.startswith(current_code + " -") or item_text.startswith(current_code + "."):
                    self.drainage_input.setCurrentIndex(i)
                    break

    def update_texture_options(self, crop_name):
        """Update texture dropdown based on selected crop from JSON requirements."""
        
        # Texture descriptions mapping
        texture_descriptions = {
            'C': 'Clay',
            'C<60s': 'Clay (<60% clay, moderately heavy)',
            'C>60s': 'Clay (>60% clay, heavy clay)',
            'C>60v': 'Clay (>60% clay, very heavy clay)',
            'CL': 'Clay Loam',
            'CSGOs': 'Clay Sandy Gravel over sand',
            'Cm': 'Clay montmorillonitic',
            'Co': 'Clay oxidic',
            'CxGOs': 'Clay over Gravel over sand',
            'CxGOv': 'Clay over Gravel oxidic/vermiculitic',
            'CxGv': 'Clay over Gravel vermiculitic',
            'L': 'Loam',
            'LS': 'Loamy Sand',
            'LcS': 'Loamy coarse Sand',
            'LfS': 'Loamy fine Sand',
            'S': 'Sand',
            'SC': 'Sandy Clay',
            'SCL': 'Sandy Clay Loam',
            'SL': 'Sandy Loam',
            'SiC': 'Silty Clay',
            'SiCL': 'Silty Clay Loam',
            'SiCm': 'Silty Clay montmorillonitic',
            'SiCs': 'Silty Clay smectitic',
            'SiL': 'Silt Loam',
            'cS': 'coarse Sand',
            'fS': 'fine Sand'
        }
        
        texture_options = {
            'Arabica Coffee': [
                'C<60s', 'C>60s', 'C>60v', 'CL', 'Cm', 'Co', 'LcS', 'LfS', 
                'S', 'SC', 'SCL', 'SL', 'SiC', 'SiCL', 'SiCm', 'cS', 'fS'
            ],
            'Banana': [
                'CL', 'C>60s', 'Cm', 'Co', 'C<60s', 'C<60v', 'C>60v', 'L', 'LS', 
                'LfS', 'S', 'SC', 'SCL', 'SL', 'SiCL', 'SiCm', 'SiCs', 'SiL', 'cS', 'fS'
            ],
            'Cocoa': [
                'C<60s', 'Co', 'SiCL', 'CL', 'SiL', 'C>60s', 'SC', 'C<60v', 'L', 'SCL', 
                'C>60v', 'SL','LfS', 'Cm', 'SiCm', 'LS', 'cS', 'fS', 'S'
                
            ],
            'Cabbage': [
                'C<60s', 'Co', 'SiCs', 'SiCL', 'CL', 'SiL', 'Si', 
                'C>60s', 'SC', 'C<60v', 'L', 'SCL', 'C>60v', 'SL', 'LS', 'LfS', 'S',
                'LcS', 'fS', 'Cm', 'SiCm', 'cS'
            ],
            'Carrots': [
                'L', 'SL', 'LfS', 'SCL', 'LS', 'SiCL', 'SC', 'SiCs', 'S', 'fS', 'LcS',
                'C<60v', 'Co', 'C>60s', 'C<60s', 'C>60v', 'Cm', 'SiCm'
            ],
            "Robusta Coffee": [
                "C<60s", "Co", "SiCL", "CL", "SC", "C>60s", "L", "SCL", 
                "SL", "LS", "LfS", "Cm", "SiCm", "C>60v", "S", "cS"],
            "Maize": [
                "C<60s", "Co", "C<60v", "SC", "SiCL", "CL", "C>60v", "SCL", 
                "C>60", "SL", "LfS", "S", "LcS", "Cm", "SiCm", "cS"],
            "Oil Palm": [
                "C<60s", "SiCa", "SiCL", "Co", "CL", "SC", "L", "SCL", "SL", 
                "LfS", "Cm", "SiCm", "cS", "fS", "S", "LcS"],
            
            "Pineapple": [
                'SCL', 'L', 'SL', 'SiL', 'SiCL', 'SC', 'LfS', 'LS', 'Co',
                'C<60v', 'C<60s', 'C>60s', 'SiC', 'fS', 'Cm', 'SiCm'
            ],
            "Sorghum": [
                'C<60s', 'SiC', 'Co', 'SiCL', 'Si', 'SiL', 'SC', 'C<60v',
                'C>60s', 'L', 'SCL', 'C>60v', 'SL', 'LfS', 'LS', 'fS', 'LcS'
            ],
            "Sugarcane": [
                'C<60s', 'SiC', 'Co', 'SiL', 'Si', 'SiCL', 'CL', 'C>60v',
                'C<60s', 'SC', 'L', 'SCL', 'C>60v', 'SL', 'LcS', 'fS', 'LS'
                'Cm', 'SiCm', 'S', 'cS'
            ],
            "Sweet Potato": [
                'C<60s', 'SiC', 'Co', 'SiCL', 'Si', 'C<60v', 'SL', 'C>60s', 'L', 'SCL', 'SC',
                'C>60v', 'LS', 'LfS', 'LcS', 'fS', 'S', 'Cm', 'SiCm'
            ]
        }
        
        # Default textures if crop not in mapping
        default_textures = [
            'C', 'CL', 'Cm', 'L', 'LS', 'S', 'SC', 'SCL', 'SL', 
            'SiC', 'SiCL', 'SiCm', 'SiL', 'cS', 'fS'
        ]
        
        if not hasattr(self, 'texture_input') or not self.texture_input:
            return
        
        # Save current selection
        current_code = self.get_texture_code()
        
        # Get texture codes for this crop
        texture_codes = texture_options.get(crop_name, default_textures)
        
        # Build dropdown items with descriptions
        items = ["Select texture..."]
        for code in sorted(texture_codes):
            description = texture_descriptions.get(code, code)
            items.append(f"{code} - {description}")
        


        # ✅ NEW: Cache the current texture options
        self._current_texture_options = items.copy()
                # Update dropdown
        self.texture_input.clear()
        self.texture_input.addItems(items)
        
        # Restore selection if valid for this crop
        if current_code and current_code in texture_codes:
            for i in range(self.texture_input.count()):
                item_text = self.texture_input.itemText(i)
                if item_text.startswith(current_code + " -"):
                    self.texture_input.setCurrentIndex(i)
                    break


    def collect_form_data(self):
        """
        Collect all form data into a dictionary for evaluation.
        
        Returns:
            dict: Complete soil and climate data matching rules_engine requirements
        """
        data = {}
        
        # ===== CLIMATE DATA =====
        data['temperature'] = self.climate_inputs['temperature'].value()
        data['rainfall'] = self.climate_inputs['rainfall'].value()
        data['humidity'] = self.climate_inputs['humidity'].value()
        
        # ===== TOPOGRAPHY =====
        data['slope'] = self.soil_inputs['slope'].value()
        
        # ===== WETNESS =====
        data['drainage'] = self.get_drainage_code()
        data['flooding'] = self.get_flooding_code()
        
        # ===== PHYSICAL SOIL =====
        data['texture'] = self.get_texture_code()
        data['soil_depth'] = self.soil_inputs['soil_depth'].value()
        data['coarse_fragments'] = self.soil_inputs['coarse_fragments'].value()
        data['caco3'] = self.soil_inputs['caco3'].value()
        data['gypsum'] = self.soil_inputs['gypsum'].value()
        
        # ===== SOIL FERTILITY =====
        data['ph'] = self.soil_inputs['ph'].value()
        data['organic_carbon'] = self.soil_inputs['organic_carbon'].value()
        data['base_saturation'] = self.soil_inputs['base_saturation'].value()
        data['sum_basic_cations'] = self.soil_inputs['sum_basic_cations'].value()  # ✅ NEW LINE
        data['cec'] = self.soil_inputs['cec'].value()
        
        # ===== SALINITY =====
        data['ec'] = self.soil_inputs['ece'].value()
        data['esp'] = self.soil_inputs['esp'].value()
        
        return data

    def validate_form_data(self):
        """
        Validate form data before evaluation.
        
        Returns:
            tuple: (is_valid, error_message)
        """
        crop_name = self.crop_input.currentText()
        if crop_name == "Select a crop...":
            return False, "⚠️ Please select a crop for evaluation"
        
        if crop_name in self.seasonal_crops:
            season_text = self.season_input.currentText()
            if season_text == "Select season...":
                return False, f"⚠️ {crop_name} is a seasonal crop. Please select a growing season."
        
        data = self.collect_form_data()
        
        if data['temperature'] <= 0:
            return False, "⚠️ Temperature must be greater than 0°C"
        if data['rainfall'] <= 0:
            return False, "⚠️ Annual rainfall must be greater than 0 mm"
        if not data['texture']:
            return False, "⚠️ Please select a soil texture"
        if not data['drainage']:
            return False, "⚠️ Please select a drainage class"
        if not data['flooding']:
            return False, "⚠️ Please select a flooding class"
        if data['ph'] <= 0 or data['ph'] > 14:
            return False, "⚠️ pH must be between 0 and 14"
        
        return True, ""

    def run_analysis(self):
        """Run complete crop suitability analysis"""
        if not self.evaluator:
            QMessageBox.critical(
                self,
                "Evaluation Engine Error",
                "The evaluation engine is not initialized.\n\n"
                "Please check that the crop requirements files are loaded correctly.\n\n"
                "Check console for details."
            )
            return
        
        is_valid, error_message = self.validate_form_data()
        if not is_valid:
            QMessageBox.warning(self, "Validation Error", error_message)
            return
        
        try:
            progress = QMessageBox(self)
            progress.setWindowTitle("Running Analysis")
            progress.setText("Evaluating crop suitability...\n\nThis may take a moment.")
            progress.setStandardButtons(QMessageBox.NoButton)
            progress.setModal(True)
            progress.show()
            QApplication.processEvents()
            
            soil_data = self.collect_form_data()
            crop_name = self.crop_input.currentText()
            
            season = None
            if crop_name in self.seasonal_crops:
                season = self.get_selected_season_code()
            
            print("\n" + "="*70)
            print("🔬 RUNNING CROP SUITABILITY EVALUATION")
            print("="*70)
            print(f"Crop: {crop_name}")
            if season:
                print(f"Season: {season}")
            print(f"\nSoil Data:")
            for key, value in soil_data.items():
                print(f"  {key}: {value}")
            print("="*70)
            
            result = self.evaluator.evaluate_suitability(
                soil_data=soil_data,
                crop_name=crop_name,
                season=season
            )

            # Save evaluation result to database
            if self.db:
                try:
                    eval_data = {
                        'input_id': getattr(self, 'current_input_id', None),
                        'crop_id': crop_name.lower().replace(' ', '_'),
                        'season': season,
                        'lsi': result['lsi'],
                        'lsc': result['lsc'],
                        'full_classification': result['full_classification'],
                        'limiting_factors': result.get('limiting_factors', ''),
                        'recommendation': ', '.join(result.get('recommendations', []))[:500],
                        'full_result': result
                    }
                    
                    eval_id = self.db.save_evaluation_result(eval_data)
                    print(f"Evaluation result saved to database (ID: {eval_id})")
                    
                except Exception as db_error:
                    print(f"Could not save evaluation to database: {db_error}")

            
            progress.close()
            progress.deleteLater()
            QApplication.processEvents()

            self.current_soil_crop = crop_name
            
            self.show_results_summary(result)
            
              # Add site_name to results for map highlighting
            selected_site = self.site_input.currentText().strip()  # ✅ FIXED: Use .currentText() for QComboBox
            if selected_site and selected_site != "Select barangay...":
                result['site_name'] = selected_site
                print(f"✅ Passing site_name to reports: '{selected_site}'")
            else:
                result['site_name'] = ''
                print("⚠️ No site name entered")
            
            self.evaluation_complete.emit(result)
            
            print("\n✅ Evaluation completed successfully")
            print("="*70 + "\n")
            
        except ValueError as e:
            progress.close()
            QMessageBox.warning(
                self,
                "Evaluation Error",
                f"Could not evaluate crop suitability:\n\n{str(e)}"
            )
            print(f"\n⚠️ Validation error: {e}\n")
            
        except Exception as e:
            progress.close()
            QMessageBox.critical(
                self,
                "Unexpected Error",
                f"An unexpected error occurred:\n\n{str(e)}\n\n"
                f"Please check the console for details."
            )
            print(f"\n❌ Error during evaluation:")
            import traceback
            traceback.print_exc()

    def show_results_summary(self, result):
        """Display a summary of evaluation results"""
        lsi = result["lsi"]
        lsc = result["lsc"]
        fullclass = result["full_classification"]
        limiting = result["limiting_factors"]
        
        # Emoji mapping
        emoji_map = {
            "S1": "",
            "S2": "",
            "S3": "",
            "N": ""
        }
        emoji = emoji_map.get(lsc, "ℹ️")
        
        # Color mapping
        color = "#2d7a2d" if lsc == "S1" else "#d4a00a" if lsc == "S2" else "#d46a0a" if lsc == "S3" else "#c0392b"
        
        # Build message
        message = f"<h2>{emoji} Evaluation Complete</h2>"
        message += f"<p><b>Crop:</b> {result['crop_name']}<br>"
        message += f"<i>{result['scientific_name']}</i></p>"
        message += f"<p><b>Land Suitability Index (LSI):</b> <span style='font-size:16pt; font-weight:bold;'>{lsi:.2f}</span></p>"
        
        # FIXED: Display classification without suffixes if S1 and no real limitations
        display_class = fullclass
        if lsc == "S1" and limiting:
            # Check if all limiting factors are at S1 level (rating 0.95 or 1.0)
            # If yes, display as pure "S1" without suffixes
            has_real_limitations = False
            if "limiting_factors_detailed" in result:
                for detail in result["limiting_factors_detailed"]:
                    if detail.get("rating", 1.0) < 0.95:  # Less than S1_1 rating
                        has_real_limitations = True
                        break
            
            # If no real limitations (all S1), show just "S1"
            if not has_real_limitations:
                display_class = lsc  # Just "S1" without suffixes
                limiting = ""  # Clear limiting factors for pure S1
        
        message += f"<p><b>Classification:</b> <span style='font-size:14pt; font-weight:bold; color:{color};'>{display_class}</span></p>"
        message += f"<p><b>Interpretation:</b><br>{result['interpretation']}</p>"
        
        # FIXED: Only show limiting factors if they exist AND classification is not pure S1
        if limiting and (lsc != "S1" or display_class != lsc):
            message += f"<p><b>⚠️ Limiting Factors:</b><br>"
            for detail in result["limiting_factors_detailed"][:3]:
                message += f"• {detail['description']} ({detail['actual_value']})<br>"
            message += "</p>"
        
        message += "<hr>"
        message += "<p style='color:#666;'>View the <b>Reports</b> tab for detailed analysis and recommendations.</p>"
        
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Crop Suitability Results")
        msgbox.setTextFormat(Qt.RichText)
        msgbox.setText(message)
        msgbox.setIcon(QMessageBox.Information)
        msgbox.setStandardButtons(QMessageBox.Ok)
        msgbox.setMinimumWidth(500)
        msgbox.exec()


    def clear_form(self):
        """Clear all form inputs"""
        self.site_input.clear()
        
        for spinbox in self.soil_inputs.values():
            spinbox.setValue(0)
        
        if hasattr(self, 'texture_input'):
            self.texture_input.setCurrentIndex(0)
        if hasattr(self, 'flooding_input'):
            self.flooding_input.setCurrentIndex(0)
        if hasattr(self, 'drainage_input'):
            self.drainage_input.setCurrentIndex(0)
        
        if self.crop_input:
            self.crop_input.setCurrentIndex(0)
        if self.season_input:
            self.season_input.setCurrentIndex(0)
        
        for spinbox in self.climate_inputs.values():
            spinbox.setValue(spinbox.minimum())
        
        print("✅ Form cleared")

    def save_data(self):
        """Save soil data to database"""
        # Validate form first
        is_valid, error_message = self.validate_form_data()
        if not is_valid:
            QMessageBox.warning(self, "Validation Error", error_message)
            return
        
        if not self.db:
            QMessageBox.warning(
                self, 
                "Database Not Available",
                "Database is not connected. Data will not be saved."
            )
            return
        
        try:
            # Collect form data
            soil_data = self.collect_form_data()
            
            # Add metadata
            soil_data['location'] = self.site_input.currentText()
            soil_data['notes'] = f"Crop: {self.crop_input.currentText()}"
            
            # Save to database
            input_id = self.db.save_soil_input(soil_data)
            
            # Store for later use
            self.current_input_id = input_id
            
            print(f"Soil data saved to database with ID: {input_id}")
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Soil data saved successfully!\n\nInput ID: {input_id}\n\n"
                "You can now run the analysis to evaluate crop suitability."
            )
            
            self.data_saved.emit(input_id)
            
        except Exception as e:
            print(f"Error saving soil data: {e}")
            import traceback
            traceback.print_exc()
            
            QMessageBox.critical(
                self,
                "Save Error",
                f"Failed to save soil data:\n{str(e)}"
            )

    def import_excel(self):
        """Import soil data from Excel file"""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Import Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if not filename:
            return
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            # Parse data - looking for PARAMETER and VALUE columns
            data_dict = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] and row[1]:  # If both Parameter and Value exist
                    param = str(row[0]).strip()
                    value = row[1]
                    if value is not None:
                        value = str(value).strip()
                        if value:  # Only store non-empty values
                            data_dict[param] = value
            
            # Map parameters to form fields
            imported_count = 0
            
            # CROP
            if "Crop Name" in data_dict:
                crop = data_dict["Crop Name"]
                for i in range(self.crop_input.count()):
                    if crop.lower() in self.crop_input.itemText(i).lower():
                        self.crop_input.setCurrentIndex(i)
                        imported_count += 1
                        break
            
            if "Season (if applicable)" in data_dict:
                season = data_dict["Season (if applicable)"]
                for i in range(self.season_input.count()):
                    if season.lower() in self.season_input.itemText(i).lower():
                        self.season_input.setCurrentIndex(i)
                        imported_count += 1
                        break
            
            # LOCATION
            if "Site Name" in data_dict:
                site_value = data_dict["Site Name"]
                # FIXED: Restore site names from cache if dropdown is empty
                if self.site_input.count() == 0:
                    self.site_input.addItems(self._cached_site_names)
                    print("Site names restored from cache during import")
                
                # Find and select the matching site
                for i in range(self.site_input.count()):
                    if self.site_input.itemText(i) == site_value:
                        self.site_input.setCurrentIndex(i)
                        imported_count += 1
                        break

            
            # CLIMATE
            if "Average Temperature (°C)" in data_dict or "Average Temperature (C)" in data_dict:
                key = "Average Temperature (°C)" if "Average Temperature (°C)" in data_dict else "Average Temperature (C)"
                try:
                    self.climate_inputs['temperature'].setValue(float(data_dict[key]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Annual Rainfall (mm)" in data_dict:
                try:
                    self.climate_inputs['rainfall'].setValue(float(data_dict["Annual Rainfall (mm)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Humidity (%)" in data_dict:
                try:
                    self.climate_inputs['humidity'].setValue(float(data_dict["Humidity (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            # TOPOGRAPHY
            if "Slope (%)" in data_dict:
                try:
                    self.soil_inputs['slope'].setValue(float(data_dict["Slope (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            # WETNESS
            if "Flooding" in data_dict:
                code = data_dict["Flooding"].strip()
                for i in range(self.flooding_input.count()):
                    if self.flooding_input.itemText(i).startswith(code + " -"):
                        self.flooding_input.setCurrentIndex(i)
                        imported_count += 1
                        break
            
            if "Drainage" in data_dict:
                code = data_dict["Drainage"].strip()
                for i in range(self.drainage_input.count()):
                    if self.drainage_input.itemText(i).startswith(code + " -"):
                        self.drainage_input.setCurrentIndex(i)
                        imported_count += 1
                        break
            
            # PHYSICAL SOIL
            if "Texture" in data_dict:
                code = data_dict["Texture"].strip()
                for i in range(self.texture_input.count()):
                    if self.texture_input.itemText(i).startswith(code + " -"):
                        self.texture_input.setCurrentIndex(i)
                        imported_count += 1
                        break
            
            if "Coarse Fragments (vol%)" in data_dict:
                try:
                    self.soil_inputs['coarse_fragments'].setValue(float(data_dict["Coarse Fragments (vol%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Soil Depth (cm)" in data_dict:
                try:
                    self.soil_inputs['soil_depth'].setValue(float(data_dict["Soil Depth (cm)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "CaCO₃ (%)" in data_dict or "CaCO3 (%)" in data_dict:
                key = "CaCO₃ (%)" if "CaCO₃ (%)" in data_dict else "CaCO3 (%)"
                try:
                    self.soil_inputs['caco3'].setValue(float(data_dict[key]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Gypsum (%)" in data_dict:
                try:
                    self.soil_inputs['gypsum'].setValue(float(data_dict["Gypsum (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            # SOIL FERTILITY
            if "Apparent CEC (cmol/kg clay)" in data_dict:
                try:
                    self.soil_inputs['cec'].setValue(float(data_dict["Apparent CEC (cmol/kg clay)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Sum of Basic Cations (cmol/kg)" in data_dict:
                try:
                    self.soil_inputs['sum_basic_cations'].setValue(float(data_dict["Sum of Basic Cations (cmol/kg)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Base Saturation (%)" in data_dict:
                try:
                    self.soil_inputs['base_saturation'].setValue(float(data_dict["Base Saturation (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "pH (H₂O)" in data_dict or "pH (H2O)" in data_dict:
                key = "pH (H₂O)" if "pH (H₂O)" in data_dict else "pH (H2O)"
                try:
                    self.soil_inputs['ph'].setValue(float(data_dict[key]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "Organic Carbon (%)" in data_dict:
                try:
                    self.soil_inputs['organic_carbon'].setValue(float(data_dict["Organic Carbon (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            # SALINITY & ALKALINITY
            if "ECe (dS/m)" in data_dict:
                try:
                    self.soil_inputs['ece'].setValue(float(data_dict["ECe (dS/m)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            if "ESP (%)" in data_dict:
                try:
                    self.soil_inputs['esp'].setValue(float(data_dict["ESP (%)"]))
                    imported_count += 1
                except ValueError:
                    pass
            
            QMessageBox.information(
                self,
                "Import Successful",
                f"Data imported from:\n{os.path.basename(filename)}\n\n"
                f"{imported_count} fields populated successfully!"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "❌ Import Error",
                f"Could not import data from Excel:\n{str(e)}"
            )


    def export_excel(self):
        """Export current form data to Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Excel File",
            f"SoilWise_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            # Collect current form data
            data = self.collect_form_data()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Soil Data"
            
            # Define styles
            header_fill = PatternFill(start_color="7d9d7f", end_color="7d9d7f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="e8f3e8", end_color="e8f3e8", fill_type="solid")
            section_font = Font(bold=True, size=11, color="3d5a3f")
            border = Border(
                left=Side(style='thin', color='d4e4d4'),
                right=Side(style='thin', color='d4e4d4'),
                top=Side(style='thin', color='d4e4d4'),
                bottom=Side(style='thin', color='d4e4d4')
            )
            
            # Set column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:B{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = "SoilWise - Soil Data Export"
            title_cell.font = Font(bold=True, size=14, color="3d5a3f")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Export date
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Helper function
            def add_section(title, items):
                nonlocal row
                # Section header
                ws.merge_cells(f'A{row}:B{row}')
                section_cell = ws[f'A{row}']
                section_cell.value = title
                section_cell.fill = section_fill
                section_cell.font = section_font
                section_cell.border = border
                row += 1
                
                # Column headers
                ws[f'A{row}'] = "PARAMETER"
                ws[f'B{row}'] = "VALUE"
                for col in ['A', 'B']:
                    cell = ws[f'{col}{row}']
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                row += 1
                
                # Add items
                for param, value in items:
                    ws[f'A{row}'] = param
                    ws[f'B{row}'] = value
                    for col in ['A', 'B']:
                        ws[f'{col}{row}'].border = border
                    row += 1
                row += 1
            
            # CROP
            add_section("🌾 CROP SELECTION", [
                ("Crop Name", self.crop_input.currentText()),
                ("Season", self.season_input.currentText() if self.season_input.isVisible() else "N/A")
            ])
            
            # LOCATION
            add_section("📍 LOCATION", [
                ("Site Name", self.site_input.currentText().strip() if isinstance(self.site_input, QComboBox) else (self.site_input.text() or "N/A"))
            ])
            
            # CLIMATE
            add_section("☁️ CLIMATE CHARACTERISTICS", [
                ("Average Temperature (°C)", data.get('temperature', 0)),
                ("Annual Rainfall (mm)", data.get('rainfall', 0)),
                ("Humidity (%)", data.get('humidity', 0))
            ])
            
            # TOPOGRAPHY
            add_section("⛰️ TOPOGRAPHY", [
                ("Slope (%)", data.get('slope', 0))
            ])
            
            # WETNESS
            add_section("💧 WETNESS", [
                ("Flooding", data.get('flooding', 'N/A')),
                ("Drainage", data.get('drainage', 'N/A'))
            ])
            
            # PHYSICAL SOIL
            add_section("🏔️ PHYSICAL SOIL CHARACTERISTICS", [
                ("Texture", data.get('texture', 'N/A')),
                ("Coarse Fragments (vol%)", data.get('coarse_fragments', 0)),
                ("Soil Depth (cm)", data.get('soil_depth', 0)),
                ("CaCO₃ (%)", data.get('caco3', 0)),
                ("Gypsum (%)", data.get('gypsum', 0))
            ])
            
            # SOIL FERTILITY
            add_section("🌱 SOIL FERTILITY CHARACTERISTICS", [
                ("Apparent CEC (cmol/kg clay)", data.get('cec', 0)),
                ("Sum of Basic Cations (cmol/kg)", data.get('sum_basic_cations', 0)),
                ("Base Saturation (%)", data.get('base_saturation', 0)),
                ("pH (H₂O)", data.get('ph', 0)),
                ("Organic Carbon (%)", data.get('organic_carbon', 0))
            ])
            
            # SALINITY & ALKALINITY
            add_section("⚗️ SALINITY & ALKALINITY", [
                ("ECe (dS/m)", data.get('ec', 0)),
                ("ESP (%)", data.get('esp', 0))
            ])
            
            # Save file
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Export Successful",
                f"Data exported successfully!\n\n"
                f"Saved to:\n{filename}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Could not export data to Excel:\n{str(e)}"
            )


    def download_template(self):
        """Generate and download Excel template with all input fields"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save Template",
            "SoilWise_Template.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Soil Data Input"
            
            # Define styles
            header_fill = PatternFill(start_color="7d9d7f", end_color="7d9d7f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="e8f3e8", end_color="e8f3e8", fill_type="solid")
            section_font = Font(bold=True, size=11, color="3d5a3f")
            border = Border(
                left=Side(style='thin', color='d4e4d4'),
                right=Side(style='thin', color='d4e4d4'),
                top=Side(style='thin', color='d4e4d4'),
                bottom=Side(style='thin', color='d4e4d4')
            )
            
            # Set column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 50
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:C{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = "SoilWise - Soil Data Input Template"
            title_cell.font = Font(bold=True, size=14, color="3d5a3f")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            
            # Instructions
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "Instructions: Fill in the VALUE column with your data. Do not modify the PARAMETER column."
            ws[f'A{row}'].font = Font(italic=True, size=10)
            row += 2
            
            # Helper function to add section
            def add_section(title, fields):
                nonlocal row
                # Section header
                ws.merge_cells(f'A{row}:C{row}')
                section_cell = ws[f'A{row}']
                section_cell.value = title
                section_cell.fill = section_fill
                section_cell.font = section_font
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                section_cell.border = border
                row += 1
                
                # Column headers
                ws[f'A{row}'] = "PARAMETER"
                ws[f'B{row}'] = "VALUE"
                ws[f'C{row}'] = "NOTES / OPTIONS"
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                row += 1
                
                # Add fields
                for param, note in fields:
                    ws[f'A{row}'] = param
                    ws[f'B{row}'] = ""
                    ws[f'C{row}'] = note
                    for col in ['A', 'B', 'C']:
                        ws[f'{col}{row}'].border = border
                    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    ws[f'C{row}'].font = Font(size=9, italic=True)
                    row += 1
                row += 1
            
            # CROP SELECTION
            add_section("CROP SELECTION", [
                ("Crop Name", "Select: Arabica Coffee, Banana, Cabbage, Carrots, Cocoa, etc."),
                ("Season (if applicable)", "For seasonal crops: january_april, may_august, september_december")
            ])
            
            # LOCATION
            add_section("LOCATION", [
                ("Site Name", "Example: Dimalna")
            ])
            
            # CLIMATE
            add_section("CLIMATE CHARACTERISTICS", [
                ("Average Temperature (°C)", "Range: 0-50°C"),
                ("Annual Rainfall (mm)", "Range: 0-5000 mm"),
                ("Humidity (%)", "Range: 0-100%")
            ])
            
            # TOPOGRAPHY
            add_section("TOPOGRAPHY", [
                ("Slope (%)", "Range: 0-100%")
            ])
            
            # WETNESS
            add_section("WETNESS", [
                ("Flooding", "Codes: Fo, F1, F2, F3, F1+"),
                ("Drainage", "Codes: good, good_gw_over_150, moderate, imperfect, poor, etc.")
            ])
            
            # PHYSICAL SOIL
            add_section("PHYSICAL SOIL CHARACTERISTICS", [
                ("Texture", "Codes: C, CL, L, S, SC, SCL, SL, SiC, SiCL, etc."),
                ("Coarse Fragments (vol%)", "Range: 0-100%"),
                ("Soil Depth (cm)", "Range: 0-300 cm"),
                ("CaCO₃ (%)", "Range: 0-100%"),
                ("Gypsum (%)", "Range: 0-100%")
            ])
            
            # SOIL FERTILITY
            add_section("SOIL FERTILITY CHARACTERISTICS", [
                ("Apparent CEC (cmol/kg clay)", "Range: 0-200"),
                ("Sum of Basic Cations (cmol/kg)", "Range: 0-100"),
                ("Base Saturation (%)", "Range: 0-100%"),
                ("pH (H₂O)", "Range: 0-14"),
                ("Organic Carbon (%)", "Range: 0-10%")
            ])
            
            # SALINITY & ALKALINITY
            add_section("SALINITY & ALKALINITY", [
                ("ECe (dS/m)", "Range: 0-20"),
                ("ESP (%)", "Range: 0-100%")
            ])
            
            # Save file
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Success",
                f"Template downloaded successfully!\n\n"
                f"Saved to: {filename}\n\n"
                f"Fill in your data and import it back using 'Import Excel' button."
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Could not create template:\n{str(e)}"
            )



if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    window = InputPage()
    window.setWindowTitle("SoilWise - Soil Data Input")
    window.resize(900, 800)
    window.show()
    sys.exit(app.exec())

    def restore_all_dropdowns(self):
        """Restore all dropdowns from cache if they were accidentally cleared."""
        if hasattr(self, 'site_input') and self.site_input.count() == 0:
            self.site_input.addItems(self._cached_site_names)
            print("✅ Site names restored from cache")

        if hasattr(self, 'flooding_input') and self.flooding_input.count() == 0:
            self.flooding_input.addItems(self._cached_flooding_options)
            print("✅ Flooding options restored from cache")

        if hasattr(self, 'drainage_input') and self.drainage_input.count() == 0:
            self.drainage_input.addItems(self._current_drainage_options)
            print("✅ Drainage options restored from cache")

        if hasattr(self, 'texture_input') and self.texture_input.count() == 0:
            self.texture_input.addItems(self._current_texture_options)
            print("✅ Texture options restored from cache")

    def verify_dropdowns(self):
        """Verify all dropdowns have options, restore if needed."""
        issues = []

        if hasattr(self, 'site_input') and self.site_input.count() == 0:
            issues.append("Site names missing")
        if hasattr(self, 'flooding_input') and self.flooding_input.count() == 0:
            issues.append("Flooding options missing")
        if hasattr(self, 'drainage_input') and self.drainage_input.count() == 0:
            issues.append("Drainage options missing")
        if hasattr(self, 'texture_input') and self.texture_input.count() == 0:
            issues.append("Texture options missing")

        if issues:
            print(f"⚠️ Dropdown issues detected: {', '.join(issues)}")
            self.restore_all_dropdowns()
            return False
        return True

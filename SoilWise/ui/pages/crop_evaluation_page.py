from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QScrollArea,
    QFrame, QGridLayout, QGroupBox, QCheckBox, QRadioButton,
    QComboBox, QPushButton, QMessageBox, QDialog, QTableWidget,
    QTableWidgetItem, QButtonGroup, QApplication, QHeaderView,
    QFileDialog
)
from PySide6.QtCore import Qt, Signal, QDateTime
from PySide6.QtGui import QFont, QColor, QPixmap, QPainter
from PySide6.QtWidgets import QGraphicsDropShadowEffect
from PySide6.QtCharts import QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis

import json
import os
import copy  
from datetime import datetime
from pathlib import Path
from database.db_manager import get_database


# Import evaluation engine
try:
    from knowledge_base.evaluation import SuitabilityEvaluator
    EVALUATOR_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Warning: Could not import SuitabilityEvaluator: {e}")
    EVALUATOR_AVAILABLE = False

# ✅ EXCEL export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
    print("✅ Excel export available (openpyxl)")
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Warning: openpyxl not installed. Excel export disabled.")
    print("   Install with: pip install openpyxl")


class EnhancedButton(QPushButton):
    """Enhanced button with modern styling"""

    def __init__(self, text, icon="", primary=False, parent=None):
        super().__init__(parent)
        self.setText(f"{icon} {text}" if icon else text)
        self.primary = primary
        self.setCursor(Qt.PointingHandCursor)
        self.setMinimumHeight(48)
        self.apply_style()

    def apply_style(self):
        """Apply modern agricultural styling"""
        if self.primary:
            self.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #7d9d7f, stop:1 #6b8a6d);
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-size: 15px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #8ab08c, stop:1 #7d9d7f);
                }
                QPushButton:pressed {
                    background: #6b8a6d;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background: white;
                    color: #5a7a5c;
                    border: 2px solid #d4e4d4;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-size: 14px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background: #f5f9f5;
                    border-color: #7d9d7f;
                }
                QPushButton:pressed {
                    background: #eaf3ea;
                }
            """)


class CropEvaluationPage(QWidget):
    """Multi-crop comparison page with enhanced caching features"""

    comparison_complete = Signal(list)  # Emits list of results
    navigate_to_input = Signal()  # Navigate to input page

    def __init__(self, parent=None):
        super().__init__(parent)
        self.crop_checkboxes = {}
        self.selected_soil_data = None
        self.saved_data_combo = None
        self.last_soil_data = None
        self.last_crop_name = None
        self.compare_status_label = None
        self.compare_btn = None

        # ✅ ENHANCED CACHING: Track all comparison conditions
        self.last_comparison_results = None
        self.last_evaluated_season = None
        self.last_evaluated_crops = None
        self.last_soil_data_hash = None  # ✅ NEW: Hash to detect soil data changes

        self.soil_data_timestamp = None
        self.season_card = None

        # Define seasonal crops
        self.seasonal_crops = {
            "Cabbage", "Carrots", "Maize",
            "Sorghum", "Sweet Potato", "Tomato"
        }

        # ✅ Create comparison history directory
        self.history_dir = Path("data/comparison_history")
        self.history_dir.mkdir(parents=True, exist_ok=True)

        # Initialize evaluation engine
        self.evaluator = None
        if EVALUATOR_AVAILABLE:
            try:
                self.evaluator = SuitabilityEvaluator()
                print("✅ Evaluation engine initialized for crop comparison")
            except Exception as e:
                print(f"⚠️ Warning: Could not initialize evaluator: {e}")
        else:
            print("⚠️ Warning: Evaluation engine not available")

        # Initialize database
        try:
            self.db = get_database()
            print("Database connected in Crop Evaluation Page")
        except Exception as e:
            print(f"Database connection failed: {e}")
            self.db = None


        self.init_ui()

    def _hash_soil_data(self, soil_data):
        """Create a hash of soil data to detect changes"""
        if not soil_data:
            return None
        # Create a sorted tuple of items for consistent hashing
        return hash(tuple(sorted(soil_data.items())))

    def set_last_soil_data(self, soil_data, crop_name=None):
        """Set the last used soil data from Input page"""
        # ✅ ENHANCED: Check if soil data actually changed
        new_hash = self._hash_soil_data(soil_data)

        if new_hash != self.last_soil_data_hash:
            ph = soil_data.get('ph', 'N/A')
            temp = soil_data.get('temperature', 'N/A')
            self.clear_evaluation_cache(
                f"New soil data received (pH: {ph}, Temp: {temp}°C)"
            )
            self.last_soil_data_hash = new_hash

        self.last_soil_data = soil_data
        self.last_crop_name = crop_name
        self.soil_data_timestamp = QDateTime.currentDateTime()

        print(f"✅ Crop Evaluation: Received soil data (Last crop: {crop_name})")
        self.update_saved_data_display()

    def clear_evaluation_cache(self, reason=""):
        """Clear cached evaluation results when conditions change"""
        self.last_comparison_results = None
        self.last_evaluated_season = None
        self.last_evaluated_crops = None

        if reason:
            print(f"🔄 Cache cleared: {reason}")
        else:
            print("🔄 Evaluation cache cleared")

    def _can_use_cached_results(self, selected_crops, season):
        """Check if cached results can be reused"""
        if not self.last_comparison_results:
            return False

        if set(selected_crops) != set(self.last_evaluated_crops or []):
            return False

        if season != self.last_evaluated_season:
            return False

        # Check if soil data hash changed
        current_hash = self._hash_soil_data(self.last_soil_data)
        if current_hash != self.last_soil_data_hash:
            return False

        return True

    def update_saved_data_display(self):
        """Update the saved data dropdown with last used data"""
        if not hasattr(self, 'saved_data_combo') or self.saved_data_combo is None:
            return

        self.saved_data_combo.clear()

        if self.last_soil_data:
            # ✅ ENHANCED: Add timestamp info
            temp = self.last_soil_data.get('temperature', 'N/A')
            ph = self.last_soil_data.get('ph', 'N/A')
            crop_info = f" (Last: {self.last_crop_name})" if self.last_crop_name else ""

            # Add time info if available
            time_info = ""
            if self.soil_data_timestamp:
                mins_ago = self.soil_data_timestamp.secsTo(QDateTime.currentDateTime()) // 60
                if mins_ago < 60:
                    time_info = f" - {mins_ago} min ago"
                elif mins_ago < 1440:  # Less than 24 hours
                    time_info = f" - {mins_ago // 60} hrs ago"

            label = f"Last Analysis{crop_info} - pH {ph}, Temp {temp}°C{time_info}"
            self.saved_data_combo.addItem(label, self.last_soil_data)
            print(f"✅ Added last soil data to dropdown: {label}")
        else:
            self.saved_data_combo.addItem("No soil data available")

        # Update button state
        self.update_compare_button_text()

    def init_ui(self):
        """Initialize enhanced user interface"""
        self.setStyleSheet("background-color: #fafcfa;")

        # Scrollable container
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: #fafcfa; }")

        container = QWidget()
        container.setStyleSheet("background-color: #fafcfa;")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(24)

        # Page header
        header = self.create_header()
        layout.addWidget(header)

        # Step 1: Soil data source
        layout.addWidget(self.create_soil_source_card())

        # Step 2: Crop selector
        layout.addWidget(self.create_crop_selector_card())

        # Step 3: Season selector
        layout.addWidget(self.create_season_card())

        # Compare button
        layout.addWidget(self.create_compare_button())

        layout.addStretch()

        scroll.setWidget(container)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

        self.update_season_card_state()

    def create_header(self):
        """Create page header"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 16)
        layout.setSpacing(8)

        title = QLabel("Crop Evaluation")
        title.setFont(QFont("Georgia", 28, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")

        desc = QLabel(
            "Compare suitability of multiple crops for your soil conditions. "
            "Select crops to evaluate and see which ones are best suited for your land."
        )
        desc.setFont(QFont("Segoe UI", 14))
        desc.setStyleSheet("color: #6a8a6c; line-height: 1.5;")
        desc.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(desc)

        return widget

    def create_soil_source_card(self):
        """Create soil data source selection card"""
        card = QGroupBox()
        card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)

        # Title
        title_label = QLabel("◉ Step 1: Select Soil Data Source")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        layout.addWidget(title_label)

        # Info label
        info_label = QLabel(
            "Use soil data from a previous analysis on the Soil Data Input page."
        )
        info_label.setFont(QFont("Segoe UI", 12))
        info_label.setStyleSheet("color: #6a8a6c;")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # Saved data dropdown
        self.saved_data_combo = QComboBox()
        self.saved_data_combo.setMinimumHeight(44)
        self.saved_data_combo.setStyleSheet("""
            QComboBox {
                background: #f9fbf9;
                border: 2px solid #e0ede0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #3d5a3f;
            }
            QComboBox:focus {
                border-color: #7d9d7f;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 5px solid transparent;
                border-top: 6px solid #3d5a3f;
                width: 0;
                height: 0;
                margin-right: 8px;
            }
        """)
        self.saved_data_combo.currentIndexChanged.connect(self.update_compare_button_text)
        self.update_saved_data_display()
        layout.addWidget(self.saved_data_combo)

        # ✅ Add "Update Data" button
        update_btn_layout = QHBoxLayout()
        update_data_btn = EnhancedButton("↻ Update Soil Data", "")
        update_data_btn.setToolTip("Navigate to Soil Data Input page to run a new analysis")
        update_data_btn.clicked.connect(self.on_update_data_clicked)
        update_btn_layout.addWidget(update_data_btn)
        update_btn_layout.addStretch()
        layout.addLayout(update_btn_layout)

        card.setLayout(layout)
        return card

    def on_update_data_clicked(self):
        """Handle update data button click"""
        reply = QMessageBox.question(
            self,
            "Navigate to Soil Data Input?",
            "This will take you to the Soil Data Input page to run a new analysis.\n\n"
            "Your current crop selections will be cleared. Continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.navigate_to_input.emit()

    def create_crop_selector_card(self):
        """Create multi-crop selector card with seasonal labels"""
        card = QGroupBox()
        card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)

        # Title
        title_label = QLabel("⚘ Step 2: Select Crops to Compare")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("color: #3d5a3f;")
        layout.addWidget(title_label)

        # Info label
        info_label = QLabel(
            "Select multiple crops to compare their suitability (minimum 2 crops). "
            "Crops marked with ◐ require season selection."
        )
        info_label.setFont(QFont("Segoe UI", 12))
        info_label.setStyleSheet("color: #6a8a6c;")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # Scrollable crop list
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(360)
        scroll.setStyleSheet("""
            QScrollArea {
                background: #f9fbf9;
                border: 2px solid #e0ede0;
                border-radius: 8px;
            }
        """)

        crop_widget = QWidget()
        crop_widget.setStyleSheet("background: #f9fbf9;")
        crop_layout = QVBoxLayout(crop_widget)
        crop_layout.setContentsMargins(16, 16, 16, 16)
        crop_layout.setSpacing(8)

        # Get available crops
        if self.evaluator:
            crops = self.evaluator.get_available_crops()
        else:
            crops = [
                "Arabica Coffee", "Banana", "Cabbage", "Carrots",
                "Cocoa", "Maize", "Oil Palm", "Pineapple",
                "Robusta Coffee", "Sorghum", "Sugarcane", "Sweet Potato", "Tomato"
            ]

        # Create checkboxes with season labels
        for crop in sorted(crops):
            crop_container = QWidget()
            crop_container.setStyleSheet("background: transparent;")
            crop_h_layout = QHBoxLayout(crop_container)
            crop_h_layout.setContentsMargins(0, 0, 0, 0)
            crop_h_layout.setSpacing(8)

            checkbox = QCheckBox(f"  {crop}")
            checkbox.setFont(QFont("Segoe UI", 13))
            checkbox.setStyleSheet("""
                QCheckBox {
                    color: #3d5a3f;
                    spacing: 8px;
                }
                QCheckBox::indicator {
                    width: 20px;
                    height: 20px;
                    border: 2px solid #d4e4d4;
                    border-radius: 4px;
                    background: white;
                }
                QCheckBox::indicator:checked {
                    background: #7d9d7f;
                    border-color: #7d9d7f;
                }
            """)
            checkbox.stateChanged.connect(self.on_crop_selection_changed)
            crop_h_layout.addWidget(checkbox)

            if crop in self.seasonal_crops:
                season_label = QLabel("◐ Seasonal")
                season_label.setFont(QFont("Segoe UI", 10))
                season_label.setStyleSheet("""
                    color: #c87b00;
                    background: #fff8dc;
                    padding: 2px 8px;
                    border-radius: 4px;
                    font-weight: 600;
                """)
                crop_h_layout.addWidget(season_label)

            crop_h_layout.addStretch()
            crop_layout.addWidget(crop_container)
            self.crop_checkboxes[crop] = checkbox

        crop_layout.addStretch()
        scroll.setWidget(crop_widget)
        layout.addWidget(scroll)

        # Helper buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)

        select_all_btn = EnhancedButton("Select All", "")
        select_all_btn.clicked.connect(self.select_all_crops)

        clear_all_btn = EnhancedButton("Clear All", "")
        clear_all_btn.clicked.connect(self.clear_all_crops)

        seasonal_btn = EnhancedButton("Seasonal Crops", "")
        seasonal_btn.clicked.connect(
            lambda: self.select_preset(list(self.seasonal_crops))
        )

        perennial_btn = EnhancedButton("Perennial Crops", "")
        perennial_btn.clicked.connect(self.select_perennial_crops)

        btn_layout.addWidget(select_all_btn)
        btn_layout.addWidget(clear_all_btn)
        btn_layout.addWidget(seasonal_btn)
        btn_layout.addWidget(perennial_btn)
        btn_layout.addStretch()

        layout.addLayout(btn_layout)

        card.setLayout(layout)
        return card

    def select_perennial_crops(self):
        """Select all non-seasonal (perennial) crops"""
        self.clear_all_crops()
        for crop_name, checkbox in self.crop_checkboxes.items():
            if crop_name not in self.seasonal_crops:
                checkbox.setChecked(True)

    def create_season_card(self):
        """Create season selection card"""
        card = QGroupBox()
        card.setObjectName("seasonCard")
        card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 12px;
                border: 1px solid #e8f1e8;
                padding: 24px;
                margin-top: 12px;
            }
        """)

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout()
        layout.setContentsMargins(28, 32, 28, 28)
        layout.setSpacing(20)

        # Title
        title_label = QLabel("◐ Step 3: Select Growing Season (for seasonal crops only)")
        title_label.setFont(QFont("Georgia", 16, QFont.Bold))
        title_label.setStyleSheet("")
        layout.addWidget(title_label)

        # Info label
        info_text = QLabel(
            "This selection applies to all seasonal crops selected above. "
            "Perennial crops are not affected by season."
        )
        info_text.setFont(QFont("Segoe UI", 11))
        info_text.setStyleSheet("color: #6a8a6c; font-style: italic;")
        info_text.setWordWrap(True)
        layout.addWidget(info_text)

        # Season radio buttons
        self.season_group = QButtonGroup(self)
        seasons = [
            ("January - April (Dry Season)", "january_april"),
            ("May - August (Wet Season)", "may_august"),
            ("September - December (Cool Season)", "september_december")
        ]

        for season_text, season_code in seasons:
            radio = QRadioButton(season_text)
            radio.setFont(QFont("Segoe UI", 13))
            radio.setStyleSheet("")
            radio.setProperty("season_code", season_code)
            radio.toggled.connect(self.on_season_changed)
            self.season_group.addButton(radio)
            layout.addWidget(radio)

        # Set first as default
        self.season_group.buttons()[0].setChecked(True)

        card.setLayout(layout)
        self.season_card = card

        return card

    def on_season_changed(self, checked):
        """Handle season selection change"""
        if checked:  # Only trigger when a button becomes checked
            new_season = self.get_selected_season()
            if self.last_evaluated_season != new_season:
                self.clear_evaluation_cache(
                    f"Season changed from {self.last_evaluated_season} to {new_season}"
                )
                self.update_compare_button_text()
                print(f"✅ Season changed to: {new_season}")

    def update_season_card_state(self):
        """Enable Step 3 only if at least one seasonal crop is selected."""
        if not self.season_card:
            return

        selected = self.get_selected_crops()
        has_seasonal = any(c in self.seasonal_crops for c in selected)

        if not has_seasonal:
            self.season_card.setEnabled(False)
            self.season_card.setStyleSheet("""
                QGroupBox#seasonCard {
                    background: transparent;
                    border: none;
                    padding: 24px;
                    margin-top: 12px;
                }
                QGroupBox#seasonCard QLabel {
                    color: #b0b0b0;
                }
                QGroupBox#seasonCard QRadioButton {
                    color: #b0b0b0;
                }
            """)
        else:
            self.season_card.setEnabled(True)
            self.season_card.setStyleSheet("""
                QGroupBox#seasonCard {
                    background: white;
                    border-radius: 12px;
                    border: 1px solid #e8f1e8;
                    padding: 24px;
                    margin-top: 12px;
                }
                QGroupBox#seasonCard QLabel {
                    color: #3d5a3f;
                }
                QGroupBox#seasonCard QRadioButton {
                    color: #4a6a4c;
                }
            """)

    def create_compare_button(self):
        """Create comparison action button"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #e8f3e8, stop:1 #f0f7f0);
                border-radius: 12px;
                border: none;
            }
        """)

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(16)

        # Status label
        self.compare_status_label = QLabel("Select at least 2 crops to compare")
        self.compare_status_label.setFont(QFont("Segoe UI", 14))
        self.compare_status_label.setStyleSheet("color: #6a8a6c;")
        self.compare_status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.compare_status_label)

        # Compare button
        self.compare_btn = EnhancedButton("Compare Crops", primary=True)
        self.compare_btn.setMinimumHeight(64)
        self.compare_btn.setEnabled(False)
        self.compare_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #7d9d7f, stop:1 #6b8a6d);
                color: white;
                border: none;
                border-radius: 10px;
                padding: 16px 32px;
                font-size: 16px;
                font-weight: 700;
            }
            QPushButton:hover:enabled {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #8ab08c, stop:1 #7d9d7f);
            }
            QPushButton:pressed {
                background: #6b8a6d;
            }
            QPushButton:disabled {
                background: #d0d0d0;
                color: #888888;
            }
        """)
        self.compare_btn.clicked.connect(self.compare_crops)
        layout.addWidget(self.compare_btn)

        return card

    def get_selected_crops(self):
        """Get list of selected crop names"""
        selected = []
        for crop_name, checkbox in self.crop_checkboxes.items():
            if checkbox.isChecked():
                selected.append(crop_name)
        return selected

    def select_all_crops(self):
        """Select all crop checkboxes"""
        for checkbox in self.crop_checkboxes.values():
            checkbox.setChecked(True)
        self.update_compare_button_text()
        self.update_season_card_state()

    def clear_all_crops(self):
        """Clear all crop checkboxes"""
        for checkbox in self.crop_checkboxes.values():
            checkbox.setChecked(False)

    def select_preset(self, crop_names):
        """Select specific preset crops"""
        self.clear_all_crops()
        for crop_name in crop_names:
            if crop_name in self.crop_checkboxes:
                self.crop_checkboxes[crop_name].setChecked(True)

    def on_crop_selection_changed(self, state):
        """Handle crop checkbox state change"""
        # ✅ NEW: Clear cache if crop selection changes
        new_selected_crops = set(self.get_selected_crops())
        old_selected_crops = set(self.last_evaluated_crops or [])

        if new_selected_crops != old_selected_crops:
            self.clear_evaluation_cache("Crop selection changed")

        self.update_compare_button_text()
        self.update_season_card_state()

    def update_compare_button_text(self):
        """Update compare button text based on selection AND data availability"""
        if not hasattr(self, 'compare_status_label') or self.compare_status_label is None:
            return
        if not hasattr(self, 'compare_btn') or self.compare_btn is None:
            return

        count = len(self.get_selected_crops())
        has_data = self.last_soil_data is not None

        selected_crops = self.get_selected_crops()
        seasonal_count = sum(1 for c in selected_crops if c in self.seasonal_crops)
        perennial_count = count - seasonal_count

        if count == 0:
            self.compare_status_label.setText("Select at least 2 crops to compare")
            self.compare_status_label.setStyleSheet("color: #6a8a6c;")
            self.compare_btn.setText("Compare Crops")
            self.compare_btn.setEnabled(False)
        elif count == 1:
            self.compare_status_label.setText(
                "Select at least one more crop for comparison"
            )
            self.compare_status_label.setStyleSheet("color: #6a8a6c;")
            self.compare_btn.setText("Compare Crops")
            self.compare_btn.setEnabled(False)
        elif not has_data:
            self.compare_status_label.setText(
                f"{count} crops selected, but no soil data available. "
                "Please run an analysis on the Soil Data Input page first."
            )
            self.compare_status_label.setStyleSheet("color: #d46a00; font-weight: 600;")
            self.compare_btn.setText("Compare Crops")
            self.compare_btn.setEnabled(False)
        else:
            if seasonal_count > 0 and perennial_count > 0:
                status = f"Ready to compare {count} crops ({seasonal_count} seasonal, {perennial_count} perennial)"
            elif seasonal_count > 0:
                status = f"Ready to compare {seasonal_count} seasonal crops"
            else:
                status = f"Ready to compare {perennial_count} perennial crops"

            self.compare_status_label.setText(status)
            self.compare_status_label.setStyleSheet("color: #6a8a6c;")
            self.compare_btn.setText(f"Compare {count} Crops")
            self.compare_btn.setEnabled(True)

    def get_selected_season(self):
        """Get selected season code"""
        selected_button = self.season_group.checkedButton()
        if selected_button:
            return selected_button.property("season_code")
        return "january_april"

    def compare_crops(self):
        """Run multi-crop comparison with smart caching"""
        if not self.evaluator:
            QMessageBox.critical(
                self,
                "Evaluation Engine Error",
                "The evaluation engine is not initialized.\n\n"
                "Please check that the crop requirements files are loaded correctly."
            )
            return

        selected_crops = self.get_selected_crops()

        if len(selected_crops) < 2:
            QMessageBox.warning(
                self,
                "Insufficient Selection",
                "Please select at least 2 crops to compare."
            )
            return

        if not self.last_soil_data:
            QMessageBox.warning(
                self,
                "No Soil Data",
                "No soil data is available.\n\n"
                "Please run an analysis on the Soil Data Input page first."
            )
            return

        season = self.get_selected_season()

        # ✅ NEW: Check if we can reuse cached results
        if self._can_use_cached_results(selected_crops, season):
            print("=" * 70)
            print("✅ USING CACHED COMPARISON RESULTS")
            print("=" * 70)
            print(f"Crops: {', '.join(selected_crops)}")
            print(f"Season: {season}")
            print("=" * 70)
            self.show_comparison_results(self.last_comparison_results, is_cached=True)
            return

        try:
            progress = QMessageBox(self)
            progress.setWindowTitle("Comparing Crops")
            progress.setText("Evaluating multiple crops... this may take a moment.")
            progress.setStandardButtons(QMessageBox.NoButton)
            progress.setModal(True)
            progress.show()
            QApplication.processEvents()

            print("=" * 70)
            print("RUNNING MULTI-CROP COMPARISON")
            print("=" * 70)
            print(f"Selected crops: {', '.join(selected_crops)}")
            print(f"Season: {season}")
            print(f"Soil data available: {self.last_soil_data is not None}")

            results = []
            for crop_name in selected_crops:
                crop_season = season if crop_name in self.seasonal_crops else None
                print(f"\nEvaluating: {crop_name}" + (f" ({crop_season})" if crop_season else ""))

                result = self.evaluator.evaluate_suitability(
                    soil_data=self.last_soil_data,
                    crop_name=crop_name,
                    season=crop_season
                )
                results.append(result)

            progress.close()
            progress.deleteLater()

            # Sort by LSI (descending)
            results.sort(key=lambda x: x['lsi'], reverse=True)

            print("\n" + "=" * 70)
            print("COMPARISON RESULTS")
            print("=" * 70)
            for i, r in enumerate(results, 1):
                print(f"{i}. {r['crop_name']}: LSI={r['lsi']:.2f}, {r['full_classification']}")
            print("=" * 70)

            # ✅ Store in cache with deep copy
            self.last_comparison_results = copy.deepcopy(results)
            self.last_evaluated_crops = selected_crops.copy()
            self.last_evaluated_season = season

            # Save comparison history
            self.save_comparison_history(results, selected_crops, season)

            # Save comparison to database
            if self.db:
                try:
                    comparison_data = {
                        'input_id': None,
                        'season': season,
                        'crop_ids': selected_crops,
                        'results': [
                            {
                                'crop_name': r['crop_name'],
                                'lsi': r['lsi'],
                                'lsc': r['lsc'],
                                'classification': r['full_classification']
                            }
                            for r in results
                        ],
                        'notes': f"Compared {len(results)} crops"
                    }
                    
                    comparison_id = self.db.save_comparison(comparison_data)
                    print(f"Comparison saved to database (ID: {comparison_id})")
                    
                except Exception as db_error:
                    print(f"Could not save comparison: {db_error}")


            # Show comparison dialog
            self.show_comparison_results(results, is_cached=False)

        except Exception as e:
            if 'progress' in locals():
                progress.close()
            QMessageBox.critical(
                self,
                "Comparison Error",
                f"An error occurred during comparison:\n\n{str(e)}\n\n"
                "Please check the console for details."
            )
            print(f"❌ Error during comparison: {e}")
            import traceback
            traceback.print_exc()

    def save_comparison_history(self, results, selected_crops, season):
        """Save comparison to history file"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = self.history_dir / f"comparison_{timestamp}.json"

            history_data = {
                "timestamp": datetime.now().isoformat(),
                "soil_data": self.last_soil_data,
                "season": season,
                "selected_crops": selected_crops,
                "results": results
            }

            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, indent=2, ensure_ascii=False)

            print(f"✅ Comparison history saved: {filename}")

        except Exception as e:
            print(f"⚠️ Warning: Could not save comparison history: {e}")

        # --- Suffix legend for limiting factor groups ---
    # Matches evaluator/rules grouping: c, t, w, s, f, n [see evaluation engine mappings]
    _SUFFIX_MAP = {
        "c": "Climate",
        "t": "Topography",
        "w": "Wetness",
        "s": "Physical Soil",
        "f": "Soil Fertility",
        "n": "Salinity/Alkalinity",
    }

    def _suffix_description(self, codes: str) -> str:
        """Convert limiting-factor suffix codes like 'ctf' into 'Climate, Topography, Soil Fertility'."""
        if not codes:
            return ""
        seen = set()
        parts = []
        for ch in str(codes):
            if ch in self._SUFFIX_MAP and ch not in seen:
                parts.append(self._SUFFIX_MAP[ch])
                seen.add(ch)
        return ", ".join(parts)

    def _extract_suffixes_from_classification(self, full_classification: str) -> str:
        """
        From 'S2cft' -> 'cft'.
        From 'S1' -> ''.
        """
        if not full_classification:
            return ""
        return "".join([ch for ch in str(full_classification) if ch.isalpha()])

    def _top_limiting_params_text(self, result: dict, max_items: int = 2) -> str:
        """
        Build short text like: 'Soil pH=5.2; Drainage Condition=poor'
        Uses limiting_factors_detailed when available.
        """
        details = result.get("limiting_factors_detailed", []) or []
        parts = []
        for d in details[:max_items]:
            desc = d.get("description") or d.get("parameter") or "Factor"
            val = d.get("actual_value", "N/A")
            parts.append(f"{desc}={val}")
        return "; ".join(parts)

    def _soil_summary_text(self, soil_data: dict) -> str:
        """Compact soil summary for recommendations."""
        if not soil_data:
            return ""
        bits = []
        if soil_data.get("ph") not in (None, "", "N/A"):
            bits.append(f"pH {soil_data.get('ph')}")
        if soil_data.get("texture"):
            bits.append(f"Texture {soil_data.get('texture')}")
        if soil_data.get("drainage"):
            bits.append(f"Drainage {soil_data.get('drainage')}")
        if soil_data.get("flooding"):
            bits.append(f"Flooding {soil_data.get('flooding')}")
        return ", ".join(bits)


    def show_comparison_results(self, results, is_cached=False):
        """Display comparison results in a clean, minimal dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Crop Comparison Results")
        dialog.resize(1200, 800)
        dialog.setMinimumSize(1000, 700)
        
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ===== HEADER SECTION =====
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: #7d9d7f;
                border: none;
                border-bottom-left-radius: 20px;
                border-bottom-right-radius: 20px;
            }
        """)
        header_layout = QVBoxLayout(header_widget)
        header_layout.setContentsMargins(40, 30, 40, 30)
        header_layout.setSpacing(8)
        
        # Title
        title = QLabel("Crop Suitability Comparison")
        title.setFont(QFont("Georgia", 24, QFont.Bold))
        title.setStyleSheet("color: white;")
        header_layout.addWidget(title)
        
        # Subtitle
        subtitle = QLabel(f"Analyzed {len(results)} crop(s) for your soil conditions")
        subtitle.setFont(QFont("Segoe UI", 12))
        subtitle.setStyleSheet("color: rgba(255, 255, 255, 0.9);")
        header_layout.addWidget(subtitle)
        
        # Cache indicator
        if is_cached:
            cache_label = QLabel("Showing cached results")
            cache_label.setFont(QFont("Segoe UI", 10))
            cache_label.setStyleSheet("color: rgba(255, 255, 255, 0.8);")
            header_layout.addWidget(cache_label)
        
        main_layout.addWidget(header_widget)
        
        # ===== CONTENT AREA =====
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea {
                background-color: #fafafa;
                border: none;
            }
            QScrollBar:vertical {
                border: none;
                background: #e8e8e8;
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: #7d9d7f;
                border-radius: 5px;
                min-height: 20px;
            }
        """)
        
        content_widget = QWidget()
        content_widget.setStyleSheet("background-color: #fafafa;")
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(40, 30, 40, 30)
        content_layout.setSpacing(24)
        
        # ===== SUMMARY CARDS =====
        summary_layout = QHBoxLayout()
        summary_layout.setSpacing(16)

        best_crop = results[0]
        worst_crop = results[-1]
        avg_lsi = sum(r['lsi'] for r in results) / len(results)

        # Best Crop Card
        best_card = self._create_summary_card(
            "Most Suitable",
            best_crop['crop_name'],
            f"LSI: {best_crop['lsi']:.2f}",
            "#2e7d32"
        )
        summary_layout.addWidget(best_card)

        # Average LSI Card
        avg_card = self._create_summary_card(
            "Average LSI",
            f"{avg_lsi:.2f}",
            f"Across {len(results)} crops",
            "#666666"
        )
        summary_layout.addWidget(avg_card)

        # Least Suitable Card
        least_card = self._create_summary_card(
            "Least Suitable",
            worst_crop['crop_name'],
            f"LSI: {worst_crop['lsi']:.2f}",
            "#c62828"
        )
        summary_layout.addWidget(least_card)

        content_layout.addLayout(summary_layout)

        
        # ===== CHART =====
        chart_card = QGroupBox()
        chart_card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 8px;
                border: 1px solid #e0e0e0;
                padding: 20px;
            }
        """)
        chart_layout = QVBoxLayout(chart_card)
        chart_layout.setContentsMargins(20, 20, 20, 20)
        
        chart_title = QLabel("Suitability Comparison Chart")
        chart_title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        chart_title.setStyleSheet("color: #333333;")
        chart_layout.addWidget(chart_title)
        
        chart_view = self.create_comparison_chart(results)
        chart_view.setMinimumHeight(280)
        chart_view.setMaximumHeight(350)
        chart_layout.addWidget(chart_view)
        
        content_layout.addWidget(chart_card)
        
        # ===== TABLE =====
        table_card = QGroupBox()
        table_card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 8px;
                border: 1px solid #e0e0e0;
                padding: 20px;
            }
        """)
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(20, 20, 20, 20)
        
        table_header = QHBoxLayout()
        table_title = QLabel("Detailed Comparison Results")
        table_title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        table_title.setStyleSheet("color: #333333;")
        table_header.addWidget(table_title)
        table_header.addStretch()
        
        legend_label = QLabel("Hover over cells for details")
        legend_label.setFont(QFont("Segoe UI", 10))
        legend_label.setStyleSheet("color: #666666;")
        table_header.addWidget(legend_label)
        
        table_layout.addLayout(table_header)
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "RANK", "CROP NAME", "LSI SCORE", "CLASSIFICATION", "LIMITING FACTORS"
        ])
        table.setRowCount(len(results))
        table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: none;
                gridline-color: #e8e8e8;
                font-size: 13px;
                color: #333333;
            }
            QHeaderView::section {
                background: #f5f5f5;
                color: #555555;
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid #cccccc;
                font-weight: 600;
                font-size: 11px;
                letter-spacing: 0.5px;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #f5f5f5;
                color: #333333;
            }
        """)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.setShowGrid(False)
        
        # Populate table
        for row, result in enumerate(results):
            # Rank
            rank_item = QTableWidgetItem(str(row + 1))
            rank_item.setTextAlignment(Qt.AlignCenter)
            rank_item.setFont(QFont("Segoe UI", 13, QFont.Bold))
            table.setItem(row, 0, rank_item)
            
            # Crop name
            crop_item = QTableWidgetItem(result['crop_name'])
            crop_item.setFont(QFont("Segoe UI", 13, QFont.Bold if row == 0 else QFont.Normal))
            table.setItem(row, 1, crop_item)
            
            # LSI Score
            lsi_val = float(result.get('lsi', 0.0))
            lsi_item = QTableWidgetItem(f"{lsi_val:.2f}")
            lsi_item.setTextAlignment(Qt.AlignCenter)
            lsi_item.setFont(QFont("Segoe UI", 13, QFont.Bold))
            table.setItem(row, 2, lsi_item)
            
            # Classification
            full_class = result.get("full_classification", "")
            lsc = result.get("lsc", "")
            suffixes = self._extract_suffixes_from_classification(full_class)
            suffix_desc = self._suffix_description(suffixes)
            
            class_display = full_class
            
            class_item = QTableWidgetItem(class_display)
            class_item.setTextAlignment(Qt.AlignCenter)
            class_item.setFont(QFont("Segoe UI", 12))
            
            # Subtle color coding
            if lsc == "S1":
                class_item.setForeground(QColor("#2e7d32"))
            elif lsc == "S2":
                class_item.setForeground(QColor("#f57c00"))
            elif lsc == "S3":
                class_item.setForeground(QColor("#e65100"))
            else:
                class_item.setForeground(QColor("#c62828"))
            
            if suffix_desc:
                class_item.setToolTip(f"{suffix_desc}")
            
            table.setItem(row, 3, class_item)
            
            # Limiting factors
            lf_codes = result.get("limiting_factors", "")
            lf_desc = self._suffix_description(lf_codes)
            
            lf_display = lf_codes if lf_codes else "None"
            
            lf_item = QTableWidgetItem(lf_display)
            lf_item.setTextAlignment(Qt.AlignCenter)
            lf_item.setFont(QFont("Segoe UI", 12))
            
            if lf_desc:
                lf_item.setToolTip(f"{lf_desc}")
            
            table.setItem(row, 4, lf_item)
        
        # Column widths
        table.setColumnWidth(0, 80)
        table.setColumnWidth(1, 180)
        table.setColumnWidth(2, 120)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        
        # Row heights
        for row in range(len(results)):
            table.setRowHeight(row, 60)
        
        table_layout.addWidget(table)
        
        table_layout.addWidget(table)

        content_layout.addWidget(table_card)

        # Suffix legend - positioned OUTSIDE the table card to prevent overlap
        legend = QLabel(
            "Suffix codes: c = Climate, t = Topography, w = Wetness, "
            "s = Physical Soil, f = Soil Fertility, n = Salinity/Alkalinity"
        )
        legend.setFont(QFont("Segoe UI", 9))
        legend.setStyleSheet("""
            color: #888888; 
            padding: 12px 20px; 
            background: white;
            border-radius: 8px;
            border: 1px solid #e0e0e0;
        """)
        legend.setWordWrap(True)
        content_layout.addWidget(legend)

        
        # ===== RECOMMENDATIONS =====
        rec_group = QGroupBox("Expert Recommendations & Analysis")
        rec_group.setFont(QFont("Segoe UI", 14, QFont.Bold))
        rec_group.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 8px;
                border: 1px solid #e0e0e0;
                padding: 20px;
                margin-top: 10px;
                font-weight: bold;
                color: #333333;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                background: white;
            }
        """)
        
        rec_layout = QVBoxLayout()
        rec_layout.setContentsMargins(20, 25, 20, 20)
        rec_layout.setSpacing(16)
        
        # Build recommendation content
        best_crop = results[0]
        best_name = best_crop.get("crop_name", "Selected crop")
        best_lsi = float(best_crop.get("lsi", 0.0))
        best_fullclass = best_crop.get("full_classification", "")
        best_lf_codes = best_crop.get("limiting_factors", "")
        best_lf_desc = self._suffix_description(best_lf_codes)
        
        soil_data = best_crop.get("soil_data") or self.last_soil_data or {}
        soil_summary = self._soil_summary_text(soil_data)
        best_limiting_params = self._top_limiting_params_text(best_crop, max_items=2)
        
        # Recommended crop
        rec_header = QLabel(f"Recommended: {best_name}")
        rec_header.setFont(QFont("Georgia", 16, QFont.Bold))
        rec_header.setStyleSheet("color: #2e7d32; padding: 10px;")
        rec_layout.addWidget(rec_header)
        
        # Why this crop
        why_best_label = QLabel("Why this crop is most suitable:")
        why_best_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        why_best_label.setStyleSheet("color: #333333;")
        rec_layout.addWidget(why_best_label)
        
        why_best_lines = []
        why_best_lines.append(f"Highest suitability score among all evaluated crops (LSI: {best_lsi:.2f})")
        if best_fullclass:
            why_best_lines.append(f"Classification: {best_fullclass}")
        if soil_summary:
            why_best_lines.append(f"Soil conditions: {soil_summary}")
        if best_lf_desc:
            why_best_lines.append(f"Main limitations: {best_lf_desc}")
        else:
            why_best_lines.append(f"No significant limitations detected")
        if best_limiting_params:
            why_best_lines.append(f"Key factors to monitor: {best_limiting_params}")
        
        why_best_text = QLabel("\n".join(f"• {line}" for line in why_best_lines))
        why_best_text.setFont(QFont("Segoe UI", 11))
        why_best_text.setStyleSheet("color: #555555; padding: 15px; line-height: 1.6;")
        why_best_text.setWordWrap(True)
        rec_layout.addWidget(why_best_text)
        
        # Other crops
        if len(results) > 1:
            rec_layout.addSpacing(10)
            
            why_others_label = QLabel("Analysis of other crops:")
            why_others_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
            why_others_label.setStyleSheet("color: #333333;")
            rec_layout.addWidget(why_others_label)
            
            other_scroll = QScrollArea()
            other_scroll.setMaximumHeight(200)
            other_scroll.setWidgetResizable(True)
            other_scroll.setFrameShape(QFrame.NoFrame)
            other_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
            
            other_widget = QWidget()
            other_layout = QVBoxLayout(other_widget)
            other_layout.setContentsMargins(0, 0, 0, 0)
            other_layout.setSpacing(10)
            
            for r in results[1:]:
                name = r.get("crop_name", "Other crop")
                lsi = float(r.get("lsi", 0.0))
                fullclass = r.get("full_classification", "")
                lsc = r.get("lsc", "")
                lf_codes = r.get("limiting_factors", "")
                lf_desc = self._suffix_description(lf_codes)
                limiting_params = self._top_limiting_params_text(r, max_items=1)
                
                crop_card = QLabel()
                crop_card.setStyleSheet("background: #fafafa; padding: 12px; border-radius: 4px;")
                
                reasons = [f"{name} (LSI: {lsi:.2f}, {fullclass})"]
                
                if lsc == "N":
                    reasons.append("Not suitable for current conditions")
                elif lsc == "S3":
                    reasons.append("Marginal - requires significant management")
                elif lsc == "S2":
                    reasons.append("Moderate - needs interventions")
                
                if lf_desc:
                    reasons.append(f"Limitations: {lf_desc}")
                if limiting_params:
                    reasons.append(f"Key issue: {limiting_params}")
                
                crop_card.setText("\n".join(f"  {line}" for line in reasons))
                crop_card.setFont(QFont("Segoe UI", 10))
                crop_card.setStyleSheet("color: #555555; background: #fafafa; padding: 12px; border-radius: 4px;")
                crop_card.setWordWrap(True)
                other_layout.addWidget(crop_card)
            
            other_scroll.setWidget(other_widget)
            rec_layout.addWidget(other_scroll)
        
        rec_group.setLayout(rec_layout)
        content_layout.addWidget(rec_group)
        
        content_layout.addStretch()
        scroll.setWidget(content_widget)
        main_layout.addWidget(scroll)
        
        # ===== FOOTER =====
        footer_widget = QWidget()
        footer_widget.setStyleSheet("QWidget { background: white; border-top: 1px solid #e0e0e0; }")
        footer_layout = QHBoxLayout(footer_widget)
        footer_layout.setContentsMargins(40, 20, 40, 20)
        footer_layout.setSpacing(12)
        
        if EXCEL_AVAILABLE:
            export_btn = EnhancedButton("Export to Excel", primary=False)
            export_btn.clicked.connect(lambda: self.export_comparison_excel(results, dialog))
            footer_layout.addWidget(export_btn)
        
        footer_layout.addStretch()
        
        close_btn = EnhancedButton("Close", primary=True)
        close_btn.setMinimumWidth(150)
        close_btn.clicked.connect(dialog.accept)
        footer_layout.addWidget(close_btn)
        
        main_layout.addWidget(footer_widget)
        
        dialog.setLayout(main_layout)
        dialog.exec()


    def _create_summary_card(self, title, main_text, sub_text, text_color):
        """Create a minimal summary card widget"""
        card = QGroupBox()
        card.setStyleSheet("""
            QGroupBox {
                background: white;
                border-radius: 8px;
                border: 1px solid #e0e0e0;
                padding: 20px;
            }
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 10))
        title_label.setStyleSheet("color: #666666; font-weight: 600;")
        layout.addWidget(title_label)
        
        main_label = QLabel(main_text)
        main_label.setFont(QFont("Segoe UI", 20, QFont.Bold))
        main_label.setStyleSheet(f"color: {text_color};")
        main_label.setWordWrap(True)
        layout.addWidget(main_label)
        
        sub_label = QLabel(sub_text)
        sub_label.setFont(QFont("Segoe UI", 9))
        sub_label.setStyleSheet("color: #888888;")
        layout.addWidget(sub_label)
        
        return card


    def create_comparison_chart(self, results):
        """Create bar chart comparing LSI values"""
        # Create bar set
        bar_set = QBarSet("Land Suitability Index (LSI)")
        bar_set.setColor(QColor("#7d9d7f"))

        categories = []
        for result in results:
            bar_set.append(result['lsi'])
            categories.append(result['crop_name'])

        # Create series
        series = QBarSeries()
        series.append(bar_set)

        # Create chart
        chart = QChart()
        chart.addSeries(series)
        chart.setTitle("Crop Suitability Comparison")
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTitleFont(QFont("Georgia", 14, QFont.Bold))

        # Axes
        axis_x = QBarCategoryAxis()
        axis_x.append(categories)
        chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_x)

        axis_y = QValueAxis()
        axis_y.setRange(0, 100)
        axis_y.setTitleText("LSI Value")
        chart.addAxis(axis_y, Qt.AlignLeft)
        series.attachAxis(axis_y)

        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)

        # Create chart view
        chart_view = QChartView(chart)
        chart_view.setRenderHint(QPainter.Antialiasing)

        return chart_view

    def export_comparison_excel(self, results, parent_dialog):
        """Export comparison results to Excel file"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(
                parent_dialog,
                "Excel Export Unavailable",
                "openpyxl is not installed.\n\n"
                "Install with: pip install openpyxl"
            )
            return

        try:
            # Get save location
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"crop_comparison_{timestamp}.xlsx"

            filepath, _ = QFileDialog.getSaveFileName(
                parent_dialog,
                "Export Comparison as Excel",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if not filepath:
                return  # User cancelled

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Crop Comparison"

            # Define styles
            header_fill = PatternFill(start_color="3d5a3f", end_color="3d5a3f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=14, color="3d5a3f")
            border = Border(
                left=Side(style='thin', color='e0ede0'),
                right=Side(style='thin', color='e0ede0'),
                top=Side(style='thin', color='e0ede0'),
                bottom=Side(style='thin', color='e0ede0')
            )

            # Classification colors
            class_colors = {
                'S1': 'C6E0B4',  # Light green
                'S2': 'FFE699',  # Light yellow
                'S3': 'F4B084',  # Light orange
                'N': 'FF9999'    # Light red
            }

            # Title
            ws['A1'] = "Crop Suitability Comparison Report"
            ws['A1'].font = title_font
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Timestamp and soil info
            row = 3
            ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(size=10)
            row += 1
            ws[f'A{row}'] = f"Number of crops compared: {len(results)}"
            row += 1

            # Soil parameters (compact format)
            if self.last_soil_data:
                ph = self.last_soil_data.get('ph', 'N/A')
                temp = self.last_soil_data.get('temperature', 'N/A')
                ws[f'A{row}'] = f"Soil pH: {ph}, Temperature: {temp}°C"
                row += 1

            row += 1  # Blank row

            # Table headers
            headers = ["Rank", "Crop", "LSI", "Classification", "Limiting Factors"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Data rows
            for rank, result in enumerate(results, 1):
                row += 1

                # Rank
                cell = ws.cell(row=row, column=1, value=rank)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

                # Crop
                cell = ws.cell(row=row, column=2, value=result['crop_name'])
                cell.border = border

                # LSI
                cell = ws.cell(row=row, column=3, value=f"{result['lsi']:.2f}")
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

                # Classification with color
                lsc = result['lsc']
                class_text = result['full_classification']
                cell = ws.cell(row=row, column=4, value=class_text)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if lsc in class_colors:
                    cell.fill = PatternFill(start_color=class_colors[lsc],
                                          end_color=class_colors[lsc],
                                          fill_type="solid")

                # Limiting factors
                limiting = result.get('limiting_factors', '')
                cell = ws.cell(row=row, column=5, value=limiting if limiting else "-")
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

            row += 2  # Blank row

            # Recommendation
            best_crop = results[0]
            ws[f'A{row}'] = (
                f"◈ Recommendation: {best_crop['crop_name']} is most suitable "
                f"(LSI: {best_crop['lsi']:.2f}, {best_crop['full_classification']})"
            )
            ws[f'A{row}'].font = Font(bold=True, color="3d5a3f", size=11)
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 30

            # Save file
            wb.save(filepath)

            QMessageBox.information(
                parent_dialog,
                "Export Successful",
                f"Comparison report exported to:\n{filepath}"
            )

            print(f"✅ Excel export successful: {filepath}")

        except Exception as e:
            QMessageBox.critical(
                parent_dialog,
                "Export Error",
                f"Failed to export Excel file:\n\n{str(e)}"
            )
            print(f"❌ Excel export error: {e}")
            import traceback
            traceback.print_exc()
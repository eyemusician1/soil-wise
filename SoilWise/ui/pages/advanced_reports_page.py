"""
SoilWise/ui/pages/reports_page.py
Simplified Reports & Analysis Page - With Collapsible Recommendations
"""

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QScrollArea, QFrame, QGridLayout, QGroupBox,
                               QPushButton, QTabWidget, QFileDialog, QMessageBox,
                               QProgressDialog)
from PySide6.QtCore import Qt, Signal, QTimer, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QFont, QColor
from PySide6.QtWidgets import QGraphicsDropShadowEffect
from datetime import datetime
import json
import openpyxl
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# For PDF export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class CollapsibleSection(QWidget):
    """Collapsible section widget with smooth animation"""
    
    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self.is_expanded = True
        self.init_ui(title)
    
    def init_ui(self, title: str):
        """Initialize the collapsible section UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Toggle button
        self.toggle_button = QPushButton()
        self.toggle_button.setCheckable(True)
        self.toggle_button.setChecked(True)
        self.toggle_button.clicked.connect(self.toggle)
        self.update_button_text(title)
        self.toggle_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #8ab08c, stop:1 #7d9d7f);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 14px 20px;
                text-align: left;
                font-size: 16px;
                font-weight: 600;
                font-family: 'Georgia';
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #9bc09d, stop:1 #8ab08c);
            }
            QPushButton:pressed {
                background: #7d9d7f;
            }
        """)
        self.toggle_button.setCursor(Qt.PointingHandCursor)
        main_layout.addWidget(self.toggle_button)
        
        # Content widget
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 12, 0, 0)
        self.content_layout.setSpacing(0)
        main_layout.addWidget(self.content_widget)
        
        # Animation
        self.animation = QPropertyAnimation(self.content_widget, b"maximumHeight")
        self.animation.setDuration(300)
        self.animation.setEasingCurve(QEasingCurve.InOutQuad)
    
    def update_button_text(self, title: str):
        """Update button text with arrow indicator"""
        arrow = "▼" if self.is_expanded else "▶"
        self.toggle_button.setText(f"{arrow} {title}")
    
    def toggle(self):
        """Toggle the section visibility"""
        self.is_expanded = not self.is_expanded
        self.update_button_text(self.toggle_button.text().split(" ", 1)[1])
        
        if self.is_expanded:
            self.expand()
        else:
            self.collapse()
    
    def collapse(self):
        """Collapse the section"""
        self.animation.setStartValue(self.content_widget.height())
        self.animation.setEndValue(0)
        self.animation.start()
    
    def expand(self):
        """Expand the section"""
        self.animation.setStartValue(self.content_widget.height())
        # Get the actual height needed
        self.content_widget.setMaximumHeight(16777215)  # Reset max height
        content_height = self.content_widget.sizeHint().height()
        self.animation.setEndValue(content_height)
        self.animation.start()
    
    def set_content(self, widget: QWidget):
        """Set the content widget"""
        self.content_layout.addWidget(widget)


class AdvancedReportsPage(QWidget):
    """Detailed Reports & Analysis Page with Collapsible Recommendations"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Reports & Analysis")
        self.resize(1100, 700)
        self.setMinimumSize(900, 600)
        self.current_results = None
        self.init_ui()
    
    def init_ui(self):
        """Initialize the reports page UI"""
        self.setStyleSheet("background-color: #fafcfa;")
        
        # Main scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: #fafcfa; }")
        
        # Container
        container = QWidget()
        container.setStyleSheet("background-color: #fafcfa;")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(24)
        
        # Page header
        header = self.create_header()
        layout.addWidget(header)
        
        # Empty state (shown when no results)
        self.empty_state = self.create_empty_state()
        layout.addWidget(self.empty_state)
        
        # Results container (hidden initially)
        self.results_container = QWidget()
        self.results_layout = QVBoxLayout(self.results_container)
        self.results_layout.setSpacing(24)
        self.results_container.setVisible(False)
        layout.addWidget(self.results_container)
        
        layout.addStretch()
        scroll.setWidget(container)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)
    
    def create_header(self):
        """Create page header"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 16)
        layout.setSpacing(8)
        
        title = QLabel("Reports & Analysis")
        title.setFont(QFont("Georgia", 28, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")
        
        desc = QLabel("Comprehensive crop suitability evaluation results and recommendations")
        desc.setFont(QFont("Segoe UI", 14))
        desc.setStyleSheet("color: #6a8a6c; line-height: 1.5;")
        desc.setWordWrap(True)
        
        layout.addWidget(title)
        layout.addWidget(desc)
        return widget
    
    def create_empty_state(self):
        """Create empty state when no results available"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 12px;
                border: none;
            }
        """)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(60, 80, 60, 80)
        layout.setSpacing(20)
        layout.setAlignment(Qt.AlignCenter)
        
        icon = QLabel("◱")
        icon.setFont(QFont("Segoe UI", 72))
        icon.setStyleSheet("color: #a8b5a8;")
        icon.setAlignment(Qt.AlignCenter)
        
        title = QLabel("No Evaluation Results Yet")
        title.setFont(QFont("Georgia", 20, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")
        title.setAlignment(Qt.AlignCenter)
        
        desc = QLabel(
            "Run a crop suitability analysis from the Soil Data Input page\n"
            "to view detailed reports and recommendations here."
        )
        desc.setFont(QFont("Segoe UI", 14))
        desc.setStyleSheet("color: #6a8a6c;")
        desc.setAlignment(Qt.AlignCenter)
        desc.setWordWrap(True)
        
        layout.addWidget(icon)
        layout.addWidget(title)
        layout.addWidget(desc)
        return card
    
    def display_results(self, results: dict):
        """Display evaluation results"""
        self.current_results = results
        
        # Hide empty state, show results
        self.empty_state.setVisible(False)
        self.results_container.setVisible(True)
        
        # Clear previous results
        while self.results_layout.count():
            child = self.results_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        
        # Add result sections
        self.results_layout.addWidget(self.create_summary_card(results))
        self.results_layout.addWidget(self.create_collapsible_recommendations(results))
        self.results_layout.addWidget(self.create_action_buttons())
    
    def create_summary_card(self, results: dict):
        """Create summary card with key metrics"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 12px;
                border: none;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(20)
        
        # Title row with timestamp
        title_row = QHBoxLayout()
        title = QLabel("◈ Evaluation Summary")
        title.setFont(QFont("Georgia", 20, QFont.Bold))
        title.setStyleSheet("color: #3d5a3f;")
        title_row.addWidget(title)
        
        timestamp = QLabel(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        timestamp.setFont(QFont("Segoe UI", 11))
        timestamp.setStyleSheet("color: #8a9a8c;")
        title_row.addStretch()
        title_row.addWidget(timestamp)
        layout.addLayout(title_row)
        
        # Metrics grid
        metrics_grid = QGridLayout()
        metrics_grid.setSpacing(20)
        
        # Crop info
        crop_widget = self.create_metric_widget(
            "Crop",
            f"{results['crop_name']}",
            f"{results['scientific_name']}"
        )
        metrics_grid.addWidget(crop_widget, 0, 0)
        
        # LSI with color coding
        lsi_color = self.get_lsi_color(results['lsc'])
        lsi_widget = self.create_metric_widget(
            "Land Suitability Index",
            f"{results['lsi']:.2f}",
            "Out of 100",
            lsi_color
        )
        metrics_grid.addWidget(lsi_widget, 0, 1)
        
        # Classification
        limiting_factors = results.get('limiting_factors', '')
        if not limiting_factors or not results.get('limiting_factors_detailed'):
            full_classification = results['lsc']
        else:
            full_classification = results['full_classification']
        
        class_text = self.get_classification_text(results['lsc'])
        class_widget = self.create_metric_widget(
            "Classification",
            full_classification,
            class_text,
            lsi_color
        )
        metrics_grid.addWidget(class_widget, 0, 2)
        
        # Limiting factors
        if not limiting_factors or not results.get('limiting_factors_detailed'):
            limiting_display = "None"
            limiting_subtitle = "No significant limitations"
            limiting_color = "#2d7a2d"
        else:
            limiting_display = limiting_factors.upper()
            limiting_subtitle = self.decode_limiting_factors(limiting_factors)
            limiting_color = "#c87b00"
        
        limiting_widget = self.create_metric_widget(
            "Limiting Factors",
            limiting_display,
            limiting_subtitle,
            limiting_color
        )
        metrics_grid.addWidget(limiting_widget, 0, 3)
        
        layout.addLayout(metrics_grid)
        
        # Season info if applicable
        if results.get('season'):
            season_container = QFrame()
            season_container.setStyleSheet("""
                QFrame {
                    background: #f9fbf9;
                    border-radius: 6px;
                    padding: 12px 16px;
                }
            """)
            season_layout = QHBoxLayout(season_container)
            season_layout.setContentsMargins(0, 0, 0, 0)
            season_label = QLabel(f"◆ Growing Season: {self.format_season(results['season'])}")
            season_label.setFont(QFont("Segoe UI", 12, QFont.DemiBold))
            season_label.setStyleSheet("color: #5a7a5c;")
            season_layout.addWidget(season_label)
            layout.addWidget(season_container)
        
        # Interpretation section
        interp_container = QFrame()
        interp_container.setStyleSheet("""
            QFrame {
                background: #f9fbf9;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        interp_layout = QVBoxLayout(interp_container)
        interp_layout.setSpacing(10)
        
        interp_label = QLabel("◉ Interpretation")
        interp_label.setFont(QFont("Segoe UI", 13, QFont.DemiBold))
        interp_label.setStyleSheet("color: #4a6a4c;")
        
        interp_text = QLabel(results['interpretation'])
        interp_text.setFont(QFont("Segoe UI", 13))
        interp_text.setStyleSheet("color: #5a7a5c; line-height: 1.6;")
        interp_text.setWordWrap(True)
        
        interp_layout.addWidget(interp_label)
        interp_layout.addWidget(interp_text)
        layout.addWidget(interp_container)
        
        # Notes if available
        if results.get('notes'):
            notes_container = QFrame()
            notes_container.setStyleSheet("""
                QFrame {
                    background: #fffef0;
                    border-radius: 6px;
                    padding: 12px 16px;
                }
            """)
            notes_layout = QVBoxLayout(notes_container)
            notes_layout.setSpacing(8)
            
            notes_label = QLabel("◈ Additional Notes")
            notes_label.setFont(QFont("Segoe UI", 11, QFont.DemiBold))
            notes_label.setStyleSheet("color: #6a5a00;")
            
            notes_text = QLabel(results['notes'])
            notes_text.setFont(QFont("Segoe UI", 11))
            notes_text.setStyleSheet("color: #6a7a6c; font-style: italic;")
            notes_text.setWordWrap(True)
            
            notes_layout.addWidget(notes_label)
            notes_layout.addWidget(notes_text)
            layout.addWidget(notes_container)
        
        return card
    
    def create_metric_widget(self, label: str, value: str, subtitle: str, color: str = "#3d5a3f"):
        """Create a metric display widget"""
        widget = QFrame()
        widget.setStyleSheet("""
            QFrame {
                background: #f9fbf9;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        
        layout = QVBoxLayout(widget)
        layout.setSpacing(8)
        
        label_widget = QLabel(label)
        label_widget.setFont(QFont("Segoe UI", 11))
        label_widget.setStyleSheet("color: #6a8a6c;")
        
        value_widget = QLabel(value)
        value_widget.setFont(QFont("Georgia", 20, QFont.Bold))
        value_widget.setStyleSheet(f"color: {color};")
        value_widget.setWordWrap(True)
        
        subtitle_widget = QLabel(subtitle)
        subtitle_widget.setFont(QFont("Segoe UI", 10))
        subtitle_widget.setStyleSheet("color: #8a9a8c;")
        subtitle_widget.setWordWrap(True)
        
        layout.addWidget(label_widget)
        layout.addWidget(value_widget)
        layout.addWidget(subtitle_widget)
        return widget
    
    def create_collapsible_recommendations(self, results: dict):
        """Create collapsible recommendations section"""
        collapsible = CollapsibleSection("Recommendations")
        recommendations_content = self.create_recommendations_content(results)
        collapsible.set_content(recommendations_content)
        return collapsible
    
    def create_recommendations_content(self, results: dict):
        """Create the recommendations content widget"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 12px;
                border: none;
            }
        """)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 15))
        card.setGraphicsEffect(shadow)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(18)
        
        subtitle = QLabel(
            "Based on the evaluation results, here are specific recommendations "
            "to optimize crop production:"
        )
        subtitle.setFont(QFont("Segoe UI", 12))
        subtitle.setStyleSheet("color: #6a8a6c; margin-bottom: 8px;")
        subtitle.setWordWrap(True)
        layout.addWidget(subtitle)
        
        # Add each recommendation
        for i, rec in enumerate(results['recommendations'], 1):
            rec_widget = QFrame()
            rec_widget.setStyleSheet("""
                QFrame {
                    background: #f9fbf9;
                    border-radius: 6px;
                    padding: 14px;
                }
            """)
            
            rec_layout = QHBoxLayout(rec_widget)
            rec_layout.setSpacing(12)
            
            # Number badge
            badge = QLabel(str(i))
            badge.setFont(QFont("Segoe UI", 12, QFont.Bold))
            badge.setStyleSheet("""
                background: #7d9d7f;
                color: white;
                border-radius: 14px;
                padding: 4px;
            """)
            badge.setAlignment(Qt.AlignCenter)
            badge.setFixedSize(28, 28)
            
            rec_text = QLabel(rec)
            rec_text.setFont(QFont("Segoe UI", 12))
            rec_text.setStyleSheet("color: #4a6a4c;")
            rec_text.setWordWrap(True)
            
            rec_layout.addWidget(badge)
            rec_layout.addWidget(rec_text, 1)
            
            layout.addWidget(rec_widget)
        
        return card
    
    def create_action_buttons(self):
        """Create action buttons (PDF / Excel only)."""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setSpacing(16)
        
        btn_export_pdf = QPushButton("Export as PDF")
        btn_export_pdf.setMinimumHeight(48)
        btn_export_pdf.setStyleSheet(self.get_secondary_button_style())
        btn_export_pdf.setCursor(Qt.PointingHandCursor)
        btn_export_pdf.clicked.connect(self.export_pdf)
        
        btn_export_excel = QPushButton("Export as Excel")
        btn_export_excel.setMinimumHeight(48)
        btn_export_excel.setStyleSheet(self.get_secondary_button_style())
        btn_export_excel.setCursor(Qt.PointingHandCursor)
        btn_export_excel.clicked.connect(self.export_excel)
        
        layout.addStretch()
        layout.addWidget(btn_export_pdf)
        layout.addWidget(btn_export_excel)
        return widget
    
    def get_secondary_button_style(self):
        """Get style for secondary buttons"""
        return """
            QPushButton {
                background: white;
                color: #5a7a5c;
                border: 2px solid #7d9d7f;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: #f5f9f5;
                border-color: #6b8a6d;
            }
            QPushButton:pressed {
                background: #eaf3ea;
            }
        """
    
    # Helper methods
    def get_lsi_color(self, lsc: str) -> str:
        """Get color based on LSC"""
        colors = {'S1': '#2d7a2d', 'S2': '#d4a00a', 'S3': '#d46a0a', 'N': '#c0392b'}
        return colors.get(lsc, '#3d5a3f')
    
    def get_classification_emoji(self, lsc: str) -> str:
        """Get emoji for classification"""
        emojis = {'S1': '🟢', 'S2': '🟡', 'S3': '🟠', 'N': '🔴'}
        return emojis.get(lsc, '⚪')
    
    def get_classification_text(self, lsc: str) -> str:
        """Get text description for classification"""
        texts = {
            'S1': 'Highly Suitable',
            'S2': 'Moderately Suitable',
            'S3': 'Marginally Suitable',
            'N': 'Not Suitable'
        }
        return texts.get(lsc, 'Unknown')
    
    def decode_limiting_factors(self, factors: str) -> str:
        """Decode limiting factor codes"""
        if not factors or factors == "None":
            return "No significant limitations"
        
        codes = {
            'c': 'Climate', 't': 'Topography', 'w': 'Wetness',
            's': 'Physical Soil', 'f': 'Soil Fertility', 'n': 'Salinity/Alkalinity'
        }
        decoded = [codes.get(f, f) for f in factors.lower()]
        return ', '.join(decoded)
    
    def format_season(self, season: str) -> str:
        """Format season code to readable text"""
        seasons = {
            'january_april': 'January - April (Dry Season)',
            'may_august': 'May - August (Wet Season)',
            'september_december': 'September - December (Cool Season)'
        }
        return seasons.get(season, season.replace('_', ' ').title())
    
    # Export methods
    def export_pdf(self):
        """Export report to PDF"""
        if not self.current_results:
            QMessageBox.warning(self, "No Results", "No evaluation results available to export.")
            return
        
        if not PDF_AVAILABLE:
            QMessageBox.warning(
                self,
                "PDF Library Not Available",
                "PDF export requires the 'reportlab' library.\n\n"
                "Install it using:\npip install reportlab\n\n"
                "Then restart the application."
            )
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Report as PDF",
            f"SoilWise_Report_{self.current_results['crop_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "PDF Files (*.pdf)"
        )
        
        if not filename:
            return
        
        try:
            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=letter,
                                topMargin=0.75*inch, bottomMargin=0.75*inch,
                                leftMargin=0.75*inch, rightMargin=0.75*inch)
            
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=rl_colors.HexColor('#3d5a3f'),
                spaceAfter=12,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=rl_colors.HexColor('#3d5a3f'),
                spaceAfter=10,
                spaceBefore=15,
                fontName='Helvetica-Bold'
            )
            
            body_style = ParagraphStyle(
                'CustomBody',
                parent=styles['BodyText'],
                fontSize=11,
                textColor=rl_colors.HexColor('#4a6a4c'),
                spaceAfter=8,
                alignment=TA_JUSTIFY,
                leading=14
            )
            
            # Title
            title = Paragraph("SoilWise - Crop Suitability Evaluation Report", title_style)
            story.append(title)
            story.append(Spacer(1, 0.1*inch))
            
            # Date
            date_text = f"<i>Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</i>"
            story.append(Paragraph(date_text, body_style))
            story.append(Spacer(1, 0.3*inch))
            
            # === EVALUATION SUMMARY ===
            story.append(Paragraph("EVALUATION SUMMARY", heading_style))
            
            summary_data = [
                ['Parameter', 'Value', 'Notes'],
                ['Crop', self.current_results['crop_name'], self.current_results['scientific_name']],
                ['LSI', f"{self.current_results['lsi']:.2f}", 'Out of 100.00'],
                ['Classification', self.current_results['full_classification'],
                self.get_classification_text(self.current_results['lsc'])],
                ['Limiting Factors',
                self.current_results.get('limiting_factors', 'None').upper() if self.current_results.get('limiting_factors') else "None",
                self.decode_limiting_factors(self.current_results.get('limiting_factors', ''))],
            ]
            
            if self.current_results.get('season'):
                summary_data.insert(2, ['Season', self.format_season(self.current_results['season']), ''])
            
            summary_table = Table(summary_data, colWidths=[2*inch, 2.5*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#7d9d7f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), rl_colors.HexColor('#f9fbf9')),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#d4e4d4')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
            
            # === INTERPRETATION ===
            story.append(Paragraph("INTERPRETATION", heading_style))
            interp_text = Paragraph(self.current_results['interpretation'], body_style)
            story.append(interp_text)
            story.append(Spacer(1, 0.2*inch))
            
            # === RECOMMENDATIONS ===
            story.append(Paragraph("RECOMMENDATIONS", heading_style))
            rec_intro = Paragraph("<i>Based on the evaluation results, here are specific recommendations:</i>", body_style)
            story.append(rec_intro)
            story.append(Spacer(1, 0.1*inch))
            
            for i, rec in enumerate(self.current_results['recommendations'], 1):
                rec_text = Paragraph(f"<b>{i}.</b> {rec}", body_style)
                story.append(rec_text)
                story.append(Spacer(1, 0.05*inch))
            
            # === LIMITING FACTORS DETAILS ===
            if self.current_results.get('limiting_factors_detailed'):
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph("LIMITING FACTORS DETAILS", heading_style))
                
                details_data = [['Factor', 'Description', 'Rating']]
                for detail in self.current_results['limiting_factors_detailed']:
                    details_data.append([
                        detail.get('factor', ''),
                        detail.get('description', ''),
                        f"{detail.get('rating', 0):.2f}"
                    ])
                
                details_table = Table(details_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
                details_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#8ab08c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), rl_colors.HexColor('#f9fbf9')),
                    ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#d4e4d4')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                
                story.append(details_table)
            
            # === FOOTER ===
            story.append(Spacer(1, 0.4*inch))
            footer_text = "<i>Report generated by SoilWise v1.0 - Crop Suitability Evaluation System</i>"
            footer_style = ParagraphStyle('Footer', parent=body_style, fontSize=9,
                                        textColor=rl_colors.HexColor('#8a9a8c'), alignment=TA_CENTER)
            story.append(Paragraph(footer_text, footer_style))
            
            # Build PDF
            doc.build(story)
            
            # Success dialog
            dialog = ExportSuccessDialog("PDF", filename, self)
            dialog.show()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Could not export PDF report:\n{str(e)}"
            )


    def export_excel(self):
        """Export report to Excel"""
        if not self.current_results:
            QMessageBox.warning(self, "No Results", "No evaluation results available to export.")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Report as Excel",
            f"SoilWise_Report_{self.current_results['crop_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Evaluation Report"
            
            # Define styles
            title_fill = PatternFill(start_color="7d9d7f", end_color="7d9d7f", fill_type="solid")
            title_font = ExcelFont(bold=True, color="FFFFFF", size=16)
            header_fill = PatternFill(start_color="8ab08c", end_color="8ab08c", fill_type="solid")
            header_font = ExcelFont(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="e8f3e8", end_color="e8f3e8", fill_type="solid")
            section_font = ExcelFont(bold=True, size=12, color="3d5a3f")
            
            # Get LSI color
            lsi_color = self.get_lsi_color(self.current_results['lsc'])
            lsi_fill = PatternFill(start_color=lsi_color.replace('#', ''),
                                   end_color=lsi_color.replace('#', ''),
                                   fill_type="solid")
            
            border = Border(
                left=Side(style='thin', color='d4e4d4'),
                right=Side(style='thin', color='d4e4d4'),
                top=Side(style='thin', color='d4e4d4'),
                bottom=Side(style='thin', color='d4e4d4')
            )
            
            # Set column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:C{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = "SoilWise - Crop Suitability Evaluation Report"
            title_cell.fill = title_fill
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            
            # Generated date
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            ws[f'A{row}'].font = ExcelFont(italic=True, size=10)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # === SUMMARY SECTION ===
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "EVALUATION SUMMARY"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].border = border
            row += 1
            
            # Headers
            ws[f'A{row}'] = "PARAMETER"
            ws[f'B{row}'] = "VALUE"
            ws[f'C{row}'] = "NOTES"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Summary data
            summary_data = [
                ("Crop", self.current_results['crop_name'], self.current_results['scientific_name']),
                ("Land Suitability Index (LSI)", f"{self.current_results['lsi']:.2f}", "Out of 100.00"),
                ("Classification", self.current_results['full_classification'],
                 self.get_classification_text(self.current_results['lsc'])),
                ("Limiting Factors",
                 self.current_results.get('limiting_factors', 'None').upper() if self.current_results.get('limiting_factors') else "None",
                 self.decode_limiting_factors(self.current_results.get('limiting_factors', ''))),
            ]
            
            if self.current_results.get('season'):
                summary_data.insert(1, ("Growing Season", self.format_season(self.current_results['season']), ""))
            
            for param, value, notes in summary_data:
                ws[f'A{row}'] = param
                ws[f'B{row}'] = value
                ws[f'C{row}'] = notes
                
                # Special formatting for LSI row
                if param == "Land Suitability Index (LSI)":
                    ws[f'B{row}'].fill = lsi_fill
                    ws[f'B{row}'].font = ExcelFont(bold=True, color="FFFFFF", size=11)
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                row += 1
            
            row += 1
            
            # === INTERPRETATION ===
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "INTERPRETATION"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].border = border
            row += 1
            
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = self.current_results['interpretation']
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws[f'A{row}'].border = border
            ws.row_dimensions[row].height = 60
            row += 2
            
            # === RECOMMENDATIONS ===
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "RECOMMENDATIONS"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].border = border
            row += 1
            
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "Based on the evaluation results, here are specific recommendations:"
            ws[f'A{row}'].font = ExcelFont(italic=True, size=10)
            row += 1
            
            for i, rec in enumerate(self.current_results['recommendations'], 1):
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = f"{i}. {rec}"
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws[f'A{row}'].border = border
                ws.row_dimensions[row].height = 30
                row += 1
            
            row += 1
            
            # === LIMITING FACTORS DETAILS ===
            if self.current_results.get('limiting_factors_detailed'):
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = "LIMITING FACTORS DETAILS"
                ws[f'A{row}'].fill = section_fill
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].border = border
                row += 1
                
                # Headers
                ws[f'A{row}'] = "FACTOR"
                ws[f'B{row}'] = "DESCRIPTION"
                ws[f'C{row}'] = "RATING"
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].font = header_font
                    ws[f'{col}{row}'].border = border
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for detail in self.current_results['limiting_factors_detailed']:
                    ws[f'A{row}'] = detail.get('factor', '')
                    ws[f'B{row}'] = detail.get('description', '')
                    ws[f'C{row}'] = f"{detail.get('rating', 0):.2f}"
                    
                    for col in ['A', 'B', 'C']:
                        ws[f'{col}{row}'].border = border
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    row += 1
            
            # === FOOTER ===
            row += 2
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "Report generated by SoilWise v1.0 - Crop Suitability Evaluation System"
            ws[f'A{row}'].font = ExcelFont(italic=True, size=9, color="8a9a8c")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            # Save file
            wb.save(filename)
            
            # Success message
            # At the end of export_excel(), replace the msg = QMessageBox... block with:
            dialog = ExportSuccessDialog("Excel", filename, self)
            dialog.show()


            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Could not export Excel report:\n{str(e)}"
            )

class ExportSuccessDialog(QWidget):
    """Clean minimalist export success dialog with fixed sizing"""
    
    def __init__(self, export_type, filename, parent=None):
        super().__init__(parent)
        self.filename = filename
        self.export_type = export_type
        self.setWindowTitle("Export")
        self.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        self.setWindowModality(Qt.ApplicationModal)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the dialog UI"""
        # Fixed size - no resizing
        self.setFixedSize(480, 250)
        
        # Set clean white background
        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
            }
        """)
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(40, 35, 40, 30)
        main_layout.setSpacing(0)  # Control spacing manually
        
        # Title
        title = QLabel("Export Successful")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title.setStyleSheet("color: #2d2d2d; background: transparent;")
        title.setFixedHeight(25)
        
        # Message
        message = QLabel(f"Your {self.export_type} report has been saved successfully.")
        message.setFont(QFont("Segoe UI", 10))
        message.setStyleSheet("color: #5a5a5a; background: transparent; margin-top: 8px;")
        message.setWordWrap(True)
        message.setFixedHeight(40)
        
        # Filename section
        file_container = QWidget()
        file_container.setStyleSheet("background: transparent;")
        file_container.setFixedHeight(65)
        file_layout = QVBoxLayout(file_container)
        file_layout.setContentsMargins(0, 15, 0, 0)
        file_layout.setSpacing(6)

        filename_text = QLabel(os.path.basename(self.filename))
        filename_text.setFont(QFont("Segoe UI", 9))
        filename_text.setStyleSheet("""
            color: #3d5a3f;
            background-color: #f8f8f8;
            padding: 8px 12px;
            border: 1px solid #e5e5e5;
            border-radius: 4px;
        """)
        filename_text.setWordWrap(True)
        filename_text.setFixedHeight(36)
        
        file_layout.addWidget(filename_text)
        
        # Spacer
        spacer = QWidget()
        spacer.setStyleSheet("background: transparent;")
        spacer.setFixedHeight(15)
        
        # Buttons
        button_container = QWidget()
        button_container.setStyleSheet("background: transparent;")
        button_container.setFixedHeight(38)
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)
        
        # Open Location button
        btn_open = QPushButton("Open Folder")
        btn_open.setFixedSize(115, 38)
        btn_open.setCursor(Qt.PointingHandCursor)
        btn_open.setFont(QFont("Segoe UI", 10))
        btn_open.setStyleSheet("""
            QPushButton {
                background: #ffffff;
                color: #5a5a5a;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 0px;
            }
            QPushButton:hover {
                background: #f8f8f8;
                border-color: #7d9d7f;
                color: #3d5a3f;
            }
            QPushButton:pressed {
                background: #efefef;
            }
        """)
        btn_open.clicked.connect(self.open_file_location)
        
        # OK button
        btn_ok = QPushButton("OK")
        btn_ok.setFixedSize(95, 38)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setFont(QFont("Segoe UI", 10, QFont.DemiBold))
        btn_ok.setStyleSheet("""
            QPushButton {
                background: #7d9d7f;
                color: #ffffff;
                border: none;
                border-radius: 4px;
                padding: 0px;
            }
            QPushButton:hover {
                background: #8ab08c;
            }
            QPushButton:pressed {
                background: #6b8a6d;
            }
        """)
        btn_ok.clicked.connect(self.close)
        
        button_layout.addStretch()
        button_layout.addWidget(btn_open)
        button_layout.addWidget(btn_ok)
        
        # Add all elements
        main_layout.addWidget(title)
        main_layout.addWidget(message)
        main_layout.addWidget(file_container)
        main_layout.addWidget(spacer)
        main_layout.addWidget(button_container)
        
        # Light shadow
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(3)
        shadow.setColor(QColor(0, 0, 0, 18))
        self.setGraphicsEffect(shadow)
    
    def open_file_location(self):
        """Open the file location in file explorer"""
        import subprocess
        import platform
        
        file_path = os.path.abspath(self.filename)
        folder_path = os.path.dirname(file_path)
        
        try:
            if platform.system() == "Windows":
                subprocess.run(['explorer', '/select,', file_path])
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(['open', '-R', file_path])
            else:  # Linux
                subprocess.run(['xdg-open', folder_path])
        except Exception as e:
            print(f"Could not open file location: {e}")
